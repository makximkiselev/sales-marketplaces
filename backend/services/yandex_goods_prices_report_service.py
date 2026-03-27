from __future__ import annotations

import csv
import io
import logging
import re
import zipfile
import asyncio
from urllib.parse import urlparse, unquote
from typing import Any

import httpx
import openpyxl
import pandas as pd

from backend.routers._shared import YANDEX_BASE_URL, _find_yandex_shop_credentials, _ym_headers
from backend.services.store_data_model import replace_yandex_goods_price_report_items

logger = logging.getLogger("uvicorn.error")


def _to_num(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        pass
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", s)
    if not match:
        return None
    try:
        return float(match.group(0))
    except Exception:
        return None


def _normalize_header(value: Any) -> str:
    s = str(value or "").strip().lower()
    s = s.replace("\xa0", " ")
    s = s.replace("*", " ")
    s = s.replace('"', " ")
    s = s.replace("«", " ")
    s = s.replace("»", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _header_index_map(headers: list[Any]) -> dict[str, int]:
    return {_normalize_header(value): idx for idx, value in enumerate(headers)}


def _pick_index(header_map: dict[str, int], *names: str) -> int | None:
    for name in names:
        key = _normalize_header(name)
        if key in header_map:
            return header_map[key]
    return None


def _looks_like_real_offer_id(value: Any) -> bool:
    raw = str(value or "").strip()
    if not raw:
        return False
    normalized = _normalize_header(raw)
    if normalized in {
        "ваш sku",
        "offer id",
        "offer_id",
        "shop sku",
        "shop_sku",
        "sku",
    }:
        return False
    if "уникальный идентификатор товара" in normalized:
        return False
    if len(raw) > 80:
        return False
    if "\n" in raw:
        return False
    # Real SKUs are short codes without sentence-like spacing.
    if " " in raw and not re.fullmatch(r"[A-Za-zА-Яа-я0-9._/()\\[\\]\\-=,\\- ]{1,80}", raw):
        return False
    if raw.count(" ") >= 3:
        return False
    return True


def _sheet_priority(title: str) -> tuple[int, str]:
    normalized = _normalize_header(title)
    if normalized == "список товаров":
        return (0, normalized)
    if normalized in {"инструкция", "описание полей", "настройки", "enums"}:
        return (2, normalized)
    return (1, normalized)


def _find_header_row_index(matrix: list[list[Any]]) -> int | None:
    if not matrix:
        return None
    for idx, row in enumerate(matrix[:25]):
        header_map = _header_index_map(row)
        offer_idx = _pick_index(header_map, "offer id", "offer_id", "shop sku", "shop_sku", "sku", "ваш sku")
        on_display_idx = _pick_index(header_map, "on_display", "on display", "цена на витрине", "на витрине")
        if offer_idx is not None and on_display_idx is not None:
            return idx
    for idx, row in enumerate(matrix[:25]):
        header_map = _header_index_map(row)
        offer_idx = _pick_index(header_map, "offer id", "offer_id", "shop sku", "shop_sku", "sku", "ваш sku")
        if offer_idx is not None:
            return idx
    return 0 if matrix else None


def _parse_rows_matrix(matrix: list[list[Any]]) -> list[dict[str, Any]]:
    if not matrix:
        return []
    header_row_index = _find_header_row_index(matrix)
    if header_row_index is None or header_row_index >= len(matrix):
        return []
    header_map = _header_index_map(matrix[header_row_index])
    offer_idx = _pick_index(header_map, "offer id", "offer_id", "shop sku", "shop_sku", "sku", "ваш sku")
    name_idx = _pick_index(header_map, "offer name", "name", "название")
    currency_idx = _pick_index(header_map, "currency", "валюта")
    shop_price_idx = _pick_index(header_map, "shop price", "ваша цена", "price", "цена магазина")
    basic_price_idx = _pick_index(header_map, "basic price", "цена до скидки", "basic_price")
    on_display_idx = _pick_index(header_map, "on_display", "on display", "цена на витрине", "на витрине")
    outside_market_idx = _pick_index(
        header_map,
        "price_value_outside_market",
        "price value outside market",
        "цена вне маркета",
    )
    on_market_idx = _pick_index(
        header_map,
        "price_value_on_market",
        "price value on market",
        "цена на маркете",
    )
    green_idx = _pick_index(header_map, "price green threshold", "green threshold", "зеленый порог цены")
    red_idx = _pick_index(header_map, "price red threshold", "red threshold", "красный порог цены")
    if offer_idx is None:
        return []
    parsed: list[dict[str, Any]] = []
    for raw in matrix[header_row_index + 1 :]:
        if offer_idx >= len(raw):
            continue
        offer_id = str(raw[offer_idx] or "").strip()
        if not _looks_like_real_offer_id(offer_id):
            continue
        on_display_raw = "" if on_display_idx is None or on_display_idx >= len(raw) else str(raw[on_display_idx] or "").strip()
        parsed.append(
            {
                "offer_id": offer_id,
                "offer_name": "" if name_idx is None or name_idx >= len(raw) else str(raw[name_idx] or "").strip(),
                "currency": "" if currency_idx is None or currency_idx >= len(raw) else str(raw[currency_idx] or "").strip(),
                "shop_price": None if shop_price_idx is None or shop_price_idx >= len(raw) else _to_num(raw[shop_price_idx]),
                "basic_price": None if basic_price_idx is None or basic_price_idx >= len(raw) else _to_num(raw[basic_price_idx]),
                "on_display_raw": on_display_raw,
                "on_display_price": _to_num(on_display_raw),
                "price_value_outside_market": None if outside_market_idx is None or outside_market_idx >= len(raw) else _to_num(raw[outside_market_idx]),
                "price_value_on_market": None if on_market_idx is None or on_market_idx >= len(raw) else _to_num(raw[on_market_idx]),
                "price_green_threshold": None if green_idx is None or green_idx >= len(raw) else _to_num(raw[green_idx]),
                "price_red_threshold": None if red_idx is None or red_idx >= len(raw) else _to_num(raw[red_idx]),
            }
        )
    return parsed


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheets = list(wb.worksheets or [])
    if wb.active and wb.active not in worksheets:
        worksheets.insert(0, wb.active)
    worksheets = sorted(worksheets, key=lambda sheet: _sheet_priority(sheet.title))
    for sheet in worksheets:
        if _sheet_priority(sheet.title)[0] >= 2:
            continue
        matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
        parsed = _parse_rows_matrix(matrix)
        if parsed:
            return parsed
    return []


def _xlsx_debug_preview(content: bytes) -> dict[str, Any]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        return {"error": str(exc)}
    preview: dict[str, Any] = {"sheets": []}
    for sheet in list(wb.worksheets or [])[:5]:
        rows: list[list[str]] = []
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            rows.append([str(cell or "").strip() for cell in list(row)[:8]])
            if idx >= 7:
                break
        preview["sheets"].append({"title": sheet.title, "rows": rows})
    return preview


def _parse_excel_via_pandas(content: bytes) -> list[dict[str, Any]]:
    try:
        sheets = pd.read_excel(io.BytesIO(content), header=None, sheet_name=None)
    except Exception:
        return []
    if isinstance(sheets, dict):
        ordered_items = sorted(sheets.items(), key=lambda item: _sheet_priority(str(item[0] or "")))
        for title, df in ordered_items:
            if _sheet_priority(str(title or ""))[0] >= 2:
                continue
            matrix = df.where(pd.notnull(df), None).values.tolist()
            parsed = _parse_rows_matrix(matrix)
            if parsed:
                return parsed
    elif hasattr(sheets, "where"):
        matrix = sheets.where(pd.notnull(sheets), None).values.tolist()
        return _parse_rows_matrix(matrix)
    return []


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    raw_text = content.decode("utf-8-sig", errors="replace")
    sample = raw_text[:4096]
    delimiters = [";", ",", "\t", "|"]
    chosen = ";"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        chosen = str(getattr(dialect, "delimiter", ";") or ";")
    except Exception:
        counts = {delim: sample.count(delim) for delim in delimiters}
        chosen = max(counts, key=counts.get) if counts else ";"
    try:
        text_stream = io.StringIO(raw_text, newline="")
        reader = csv.reader(text_stream, delimiter=chosen)
        return _parse_rows_matrix([list(row) for row in reader])
    except csv.Error:
        # Some reports contain broken line breaks inside fields. Fall back to a
        # simpler split-based parse instead of failing the whole coinvest job.
        lines = raw_text.splitlines()
        reader = csv.reader(lines, delimiter=chosen)
        return _parse_rows_matrix([list(row) for row in reader])


def _parse_html_table(content: bytes) -> list[dict[str, Any]]:
    try:
        tables = pd.read_html(io.BytesIO(content))
    except Exception:
        return []
    for table in tables:
        matrix = table.where(pd.notnull(table), None).values.tolist()
        parsed = _parse_rows_matrix(matrix)
        if parsed:
            return parsed
    return []


def _looks_like_html(content: bytes) -> bool:
    head = content[:512].decode("utf-8", errors="ignore").lower()
    return "<html" in head or "<table" in head


def _looks_like_xlsx(content: bytes) -> bool:
    return bool(content[:4] == b"PK\x03\x04")


def _guess_file_name(file_url: str, file_name: str) -> str:
    explicit = str(file_name or "").strip()
    if explicit:
        return explicit
    parsed = urlparse(str(file_url or "").strip())
    candidate = unquote(parsed.path.rsplit("/", 1)[-1]) if parsed.path else ""
    return candidate.strip()


def _parse_report_file(content: bytes, filename: str) -> list[dict[str, Any]]:
    lower = str(filename or "").strip().lower()
    if lower.endswith(".xlsx"):
        return _parse_xlsx(content) or _parse_excel_via_pandas(content)
    if lower.endswith(".xls"):
        return _parse_excel_via_pandas(content)
    if lower.endswith(".csv"):
        return _parse_csv(content)
    if lower.endswith(".html") or lower.endswith(".htm"):
        return _parse_html_table(content)
    return []


def _parse_report_bytes(content: bytes) -> list[dict[str, Any]]:
    try:
        if zipfile.is_zipfile(io.BytesIO(content)):
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                for name in zf.namelist():
                    with zf.open(name) as fh:
                        file_bytes = fh.read()
                        parsed = _parse_report_file(file_bytes, name)
                        if not parsed:
                            parsed = _parse_report_bytes(file_bytes)
                        if parsed:
                            return parsed
    except Exception:
        pass
    if _looks_like_xlsx(content):
        parsed = _parse_xlsx(content) or _parse_excel_via_pandas(content)
        if parsed:
            return parsed
    if _looks_like_html(content):
        parsed = _parse_html_table(content)
        if parsed:
            return parsed
    parsed = _parse_csv(content)
    if parsed:
        return parsed
    for name in ("report.xlsx", "report.xls", "report.csv", "report.html"):
        parsed = _parse_report_file(content, name)
        if parsed:
            return parsed
    return []


async def _generate_report(*, business_id: str, api_key: str) -> str:
    url = f"{YANDEX_BASE_URL}/reports/goods-prices/generate"
    async with httpx.AsyncClient(timeout=60) as client:
        logger.warning("[coinvest_report] generate request business_id=%s url=%s", business_id, url)
        resp = await client.post(url, headers=_ym_headers(api_key), json={"businessId": int(business_id)})
        logger.warning("[coinvest_report] generate response business_id=%s status=%s", business_id, resp.status_code)
        resp.raise_for_status()
        data = resp.json() if resp.content else {}
    report_id = str(
        ((data.get("result") or {}) if isinstance(data.get("result"), dict) else {}).get("reportId")
        or data.get("reportId")
        or ""
    ).strip()
    if not report_id:
        raise RuntimeError("Маркет не вернул reportId для goods prices report")
    return report_id


async def _wait_report_download_url(*, report_id: str, api_key: str) -> tuple[str, str]:
    url = f"{YANDEX_BASE_URL}/reports/info/{report_id}"
    last_error = ""
    async with httpx.AsyncClient(timeout=60) as client:
        for _ in range(20):
            resp = await client.get(url, headers=_ym_headers(api_key))
            if resp.status_code >= 400:
                last_error = resp.text[:400]
                resp.raise_for_status()
            data = resp.json() if resp.content else {}
            result = data.get("result") if isinstance(data.get("result"), dict) else data
            status = str(result.get("status") or data.get("status") or "").strip().upper()
            file_url = str(
                result.get("file")
                or result.get("url")
                or result.get("fileUrl")
                or result.get("downloadUrl")
                or ""
            ).strip()
            file_name = str(result.get("fileName") or result.get("filename") or "").strip()
            if file_url:
                return file_url, file_name
            if status in {"DONE", "SUCCESS", "READY"} and file_url:
                return file_url, file_name
            if status in {"FAILED", "ERROR"}:
                raise RuntimeError(f"Маркет не смог подготовить goods prices report: {data}")
            await asyncio.sleep(2)
    raise RuntimeError(f"Не удалось дождаться goods prices report: {last_error or report_id}")


async def refresh_yandex_goods_prices_report_for_store(*, store_uid: str, campaign_id: str) -> dict[str, Any]:
    creds = _find_yandex_shop_credentials(campaign_id)
    if not creds:
        raise ValueError(f"credentials_not_found:{campaign_id}")
    business_id, _campaign_id, api_key = creds
    report_id = await _generate_report(business_id=business_id, api_key=api_key)
    file_url, file_name = await _wait_report_download_url(report_id=report_id, api_key=api_key)
    resolved_file_name = _guess_file_name(file_url, file_name)
    async with httpx.AsyncClient(timeout=120) as client:
        file_resp = await client.get(file_url)
        file_resp.raise_for_status()
        rows = _parse_report_bytes(file_resp.content)
        logger.warning(
            "[coinvest_report] downloaded store_uid=%s campaign_id=%s report_id=%s file_name=%s content_type=%s bytes=%s parsed_rows=%s",
            store_uid,
            campaign_id,
            report_id,
            resolved_file_name,
            str(file_resp.headers.get("content-type") or "").strip(),
            len(file_resp.content or b""),
            len(rows),
        )
        if not rows and _looks_like_xlsx(file_resp.content):
            logger.warning(
                "[coinvest_report] xlsx_preview store_uid=%s campaign_id=%s report_id=%s preview=%s",
                store_uid,
                campaign_id,
                report_id,
                _xlsx_debug_preview(file_resp.content),
            )
    loaded = replace_yandex_goods_price_report_items(
        store_uid=store_uid,
        rows=[
            {
                **row,
                "source_updated_at": report_id,
            }
            for row in rows
        ],
    )
    logger.warning("[coinvest_report] store_uid=%s campaign_id=%s report_id=%s rows=%s", store_uid, campaign_id, report_id, loaded)
    return {"store_uid": store_uid, "campaign_id": campaign_id, "report_id": report_id, "rows": loaded, "file_name": resolved_file_name}
