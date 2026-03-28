from __future__ import annotations

import datetime
import io
import json
import re
import uuid
import xml.etree.ElementTree as ET
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

import httpx
from apscheduler.schedulers.background import BackgroundScheduler
from google.oauth2.service_account import Credentials
import openpyxl
try:
    from playwright.async_api import async_playwright  # type: ignore
except Exception:  # pragma: no cover
    async_playwright = None  # type: ignore

from backend.services.integrations import (
    get_data_flow_settings,
    get_google_credentials,
    save_scoped_data_flow_settings,
)
from backend.services.store_data_model import (
    bulk_apply_pricing_defaults,
    get_fx_rates_cache,
    get_category_tree_cache_paths,
    get_pricing_category_tree,
    get_pricing_store_settings,
    replace_pricing_category_tree,
    replace_category_tree_cache_nodes,
    replace_fx_rates_cache,
    seed_pricing_category_settings_if_null,
    upsert_pricing_store_settings,
    upsert_pricing_category_setting,
    upsert_store,
    upsert_store_dataset,
    _connect,
)
from backend.services.storage import (
    load_integrations,
    load_sources,
    save_integrations,
    save_sources,
)
from backend.services.source_tables import get_registered_source_table


def parse_sheet_id(*args, **kwargs):
    from backend.services.gsheets import parse_sheet_id as impl
    return impl(*args, **kwargs)


def read_sheet_preview(*args, **kwargs):
    from backend.services.gsheets import read_sheet_preview as impl
    return impl(*args, **kwargs)


def ensure_source_table(*args, **kwargs):
    from backend.services.source_tables import ensure_source_table as impl
    return impl(*args, **kwargs)


def replace_source_rows(*args, **kwargs):
    from backend.services.source_tables import replace_source_rows as impl
    return impl(*args, **kwargs)

scheduler = BackgroundScheduler()

YANDEX_BASE_URL = "https://api.partner.market.yandex.ru/v2"
OZON_BASE_URL = "https://api-seller.ozon.ru"
BASE_DIR = Path(__file__).resolve().parents[2]
GOOGLE_KEYS_DIR = BASE_DIR / "data" / "config" / "google_keys"
GOOGLE_KEYS_DIR.mkdir(parents=True, exist_ok=True)

GSHEETS_MAPPING_TEMPLATES = {
    "pricing_input": {
        "label": "Ценообразование: входные данные",
        "required_fields": [
            "sku_primary",
            "price_site",
            "price_fbs_gt",
            "cogs",
            "stock",
        ],
    },
    "pricing_output": {
        "label": "Ценообразование: выгрузка рассчитанных цен",
        "required_fields": [
            "sku_primary",
            "new_price",
            "boost_bid",
            "op_pct",
        ],
    },
    "sales_import": {
        "label": "Аналитика продаж: импорт",
        "required_fields": [
            "order_id",
            "sku_primary",
            "qty",
            "sale_price",
            "sale_date",
        ],
    },
    "custom": {
        "label": "Пользовательский",
        "required_fields": [],
    },
}


def _mapping_configured(src: dict) -> bool:
    mapping = src.get("mapping")
    return isinstance(mapping, dict) and len(mapping) > 0


def _normalize_source_id(value: str) -> str:
    raw = (value or "").strip().lower()
    if not raw:
        return "gsheets_source"
    raw = re.sub(r"[^a-z0-9_]+", "_", raw)
    raw = re.sub(r"_+", "_", raw).strip("_")
    return raw or "gsheets_source"


def _to_bool(v, default: bool = False) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in {"1", "true", "yes", "on"}
    return default


def _validate_service_account_payload(service_account_json: str) -> tuple[dict, str, str]:
    try:
        parsed = json.loads(service_account_json)
    except Exception as e:
        raise ValueError(f"Невалидный JSON: {e}") from e

    required_fields = [
        "type",
        "project_id",
        "private_key_id",
        "private_key",
        "client_email",
        "client_id",
        "token_uri",
    ]
    missing = [k for k in required_fields if not str(parsed.get(k) or "").strip()]
    if missing:
        raise ValueError(f"В JSON отсутствуют обязательные поля: {', '.join(missing)}")
    if str(parsed.get("type") or "").strip() != "service_account":
        raise ValueError("JSON должен быть ключом service_account")

    # Дополнительная валидация формата ключа через google-auth
    try:
        Credentials.from_service_account_info(parsed)
    except Exception as e:
        raise ValueError(f"Ключ service_account невалиден: {e}") from e

    client_email = str(parsed.get("client_email") or "").strip()
    private_key_id = str(parsed.get("private_key_id") or "").strip()
    return parsed, client_email, private_key_id


def _catalog_db_connect():
    return _connect()


def _safe_source_table_name(table_name: str) -> str:
    t = str(table_name or "").strip()
    if not re.fullmatch(r"source_items_[a-z0-9_]+", t):
        raise ValueError(f"Некорректное имя таблицы источника: {t}")
    return t


def _split_subcategory_path(value: str) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    return [p.strip() for p in raw.split("/") if str(p).strip()]


def _humanize_group_id(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    raw = re.sub(r"[_\-]+", " ", raw)
    raw = re.sub(r"(?<=[a-zа-я])(?=[A-ZА-Я])", " ", raw)
    raw = re.sub(r"(?<=[A-Za-zА-Яа-я])(?=[0-9])", " ", raw)
    raw = re.sub(r"(?<=[0-9])(?=[A-Za-zА-Яа-я])", " ", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    return raw


def _clean_brand_line(value: str, brand: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if brand:
        text = re.sub(re.escape(brand), "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(смартфон|телефон|мобильный телефон|планшет|ноутбук|смарт[- ]?часы|часы|наушники|гарнитура)\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip(" -_/")
    return text


def _extract_ozon_attribute_values(payload: dict) -> dict[int, list[str]]:
    attrs = ((payload.get("product_attrs_v4") or {}).get("attributes") or []) if isinstance(payload, dict) else []
    out: dict[int, list[str]] = {}
    for attr in attrs:
        if not isinstance(attr, dict):
            continue
        try:
            attr_id = int(attr.get("id"))
        except Exception:
            continue
        values: list[str] = []
        for node in attr.get("values") or []:
            if not isinstance(node, dict):
                continue
            value = str(node.get("value") or "").strip()
            if value:
                values.append(value)
        if values:
            out[attr_id] = values
    return out


def _clean_ozon_line_candidate(value: str, brand: str) -> str:
    text = _clean_brand_line(value, brand)
    if not text:
        return ""
    text = re.sub(r"\b(?:\d+\s?(?:gb|tb)|esim|nano\s*sim|wi[- ]?fi|lte|5g|4g|global|ростест)\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(?:silver|deep blue|cosmic orange|pink|yellow|blue|black|white|green|violet|purple|gray|grey|gold)\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(?:серебристый|синий|оранжевый|розовый|желтый|голубой|черный|белый|зеленый|фиолетовый|серый|золотой)\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip(" -_/,")
    return text


def _extract_brand_line(row: dict) -> tuple[str, str]:
    payload = row.get("payload") if isinstance(row.get("payload"), dict) else {}
    ozon_attrs = _extract_ozon_attribute_values(payload)
    if payload.get("product_attrs_v4") or ozon_attrs:
        brand = str(
            (ozon_attrs.get(85) or [None])[0]
            or (ozon_attrs.get(10315) or [None])[0]
            or ""
        ).strip()
        line_candidates = [
            _clean_ozon_line_candidate(str((ozon_attrs.get(9048) or [None])[0] or ""), brand),
            _clean_ozon_line_candidate(str((ozon_attrs.get(12141) or [None])[0] or ""), brand),
            _clean_ozon_line_candidate(str((ozon_attrs.get(22390) or [None])[0] or ""), brand),
            _clean_ozon_line_candidate(str((ozon_attrs.get(5219) or [None])[0] or ""), brand),
        ]
        line = next((candidate for candidate in line_candidates if candidate), "")
        if not line:
            line = _clean_ozon_line_candidate(str(row.get("name") or "").strip(), brand)
        if line and brand and line.lower() == brand.lower():
            line = ""
        return brand, line

    offer_mapping = payload.get("offer_mapping") if isinstance(payload.get("offer_mapping"), dict) else {}
    offer = offer_mapping.get("offer") if isinstance(offer_mapping.get("offer"), dict) else {}
    mapping = offer_mapping.get("mapping") if isinstance(offer_mapping.get("mapping"), dict) else {}

    brand = str(
        offer.get("vendor")
        or mapping.get("vendor")
        or payload.get("vendor")
        or ""
    ).strip()

    line_candidates = [
        _clean_brand_line(str(mapping.get("marketModelName") or "").strip(), brand),
        _clean_brand_line(str(offer.get("groupId") or "").strip(), brand),
        _clean_brand_line(_humanize_group_id(str(offer.get("groupId") or "").strip()), brand),
    ]
    line = next((candidate for candidate in line_candidates if candidate), "")

    if not line:
        name = str(row.get("name") or "").strip()
        if brand and name.lower().startswith(brand.lower()):
            line = name[len(brand):].strip(" -_/")
        else:
            line = name

    line = _clean_brand_line(line, brand)
    if line and line.lower() == brand.lower():
        line = ""
    return brand, line


def _catalog_path_from_row(row: dict) -> list[str]:
    category = str(row.get("category") or "").strip()
    sub = _split_subcategory_path(str(row.get("subcategory") or ""))
    brand, line = _extract_brand_line(row)
    return [x for x in [category, *sub, brand, line] if x]


def _catalog_tree_from_paths(paths: list[list[str]]) -> list[dict]:
    root: dict[str, dict] = {}
    for path in paths:
        if not path:
            continue
        cur = root
        for part in path:
            if part not in cur:
                cur[part] = {"__children__": {}}
            cur = cur[part]["__children__"]

    def _natural_desc_numbers_key(value: str):
        parts = re.split(r"(\d+)", str(value or "").lower())
        key: list[tuple[int, Any]] = []
        for part in parts:
            if not part:
                continue
            if part.isdigit():
                key.append((0, -int(part)))
            else:
                key.append((1, part))
        return key

    def build(node_map: dict[str, dict], depth: int = 0) -> list[dict]:
        names = sorted(node_map.keys(), key=_natural_desc_numbers_key)
        out = []
        for name in names:
            child_map = node_map[name].get("__children__", {})
            out.append({"name": name, "children": build(child_map, depth + 1)})
        return out

    return build(root, 0)


def _read_source_rows(table_name: str) -> list[dict]:
    t = _safe_source_table_name(table_name)
    with _catalog_db_connect() as conn:
        if hasattr(conn, "text_factory"):
            conn.text_factory = lambda b: b.decode("utf-8", "replace")
        rows = conn.execute(
            f'SELECT sku, name, category, subcategory, price, currency, updated_at, payload_json FROM "{t}"'
        ).fetchall()
    out: list[dict] = []
    for r in rows:
        payload = {}
        try:
            payload = json.loads(r["payload_json"] or "{}")
        except Exception:
            payload = {}
        out.append(
            {
                "sku": str(r["sku"] or "").strip(),
                "name": str(r["name"] or "").strip(),
                "category": str(r["category"] or "").strip(),
                "subcategory": str(r["subcategory"] or "").strip(),
                "price": r["price"],
                "currency": str(r["currency"] or "").strip(),
                "updated_at": str(r["updated_at"] or "").strip(),
                "payload": payload if isinstance(payload, dict) else {},
            }
        )
    return out


def _parse_date_ymd(value: str) -> datetime.date:
    return datetime.date.fromisoformat(str(value))


def _parse_date_ru(value: str) -> datetime.date | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(raw, fmt).date()
        except Exception:
            pass
    return None


def _parse_decimal(value) -> float | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace("\xa0", "").replace(" ", "").replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", raw)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


class _SimpleHtmlTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tables: list[dict] = []
        self._in_table = False
        self._in_row = False
        self._in_cell = False
        self._cell_tag = ""
        self._current_table: dict | None = None
        self._current_row: list[str] = []
        self._current_cell_parts: list[str] = []
        self._recent_text: list[str] = []
        self._capture_attrs: dict[str, str] = {}

    def handle_starttag(self, tag, attrs):
        attrs_map = {k: v for k, v in attrs}
        if tag == "table":
            self._in_table = True
            self._current_table = {"rows": [], "context": " ".join(self._recent_text[-8:]).strip().lower()}
        elif self._in_table and tag == "tr":
            self._in_row = True
            self._current_row = []
        elif self._in_table and self._in_row and tag in {"td", "th"}:
            self._in_cell = True
            self._cell_tag = tag
            self._capture_attrs = attrs_map
            self._current_cell_parts = []

    def handle_endtag(self, tag):
        if self._in_table and self._in_row and self._in_cell and tag in {"td", "th"}:
            text = re.sub(r"\s+", " ", "".join(self._current_cell_parts)).strip()
            self._current_row.append(text)
            self._in_cell = False
            self._cell_tag = ""
            self._capture_attrs = {}
            self._current_cell_parts = []
        elif self._in_table and self._in_row and tag == "tr":
            if self._current_table is not None and any(c.strip() for c in self._current_row):
                self._current_table["rows"].append(self._current_row)
            self._in_row = False
            self._current_row = []
        elif tag == "table" and self._in_table:
            if self._current_table and self._current_table.get("rows"):
                self.tables.append(self._current_table)
            self._in_table = False
            self._current_table = None

    def handle_data(self, data):
        txt = str(data or "")
        if not txt.strip():
            return
        if self._in_table and self._in_row and self._in_cell:
            self._current_cell_parts.append(txt)
        else:
            clean = re.sub(r"\s+", " ", txt).strip()
            if clean:
                self._recent_text.append(clean)
                if len(self._recent_text) > 50:
                    self._recent_text = self._recent_text[-50:]


async def _fetch_cbr_usd_rates(date_from: datetime.date, date_to: datetime.date) -> list[dict]:
    url = "https://www.cbr.ru/scripts/XML_dynamic.asp"
    params = {
        "date_req1": date_from.strftime("%d/%m/%Y"),
        "date_req2": date_to.strftime("%d/%m/%Y"),
        "VAL_NM_RQ": "R01235",  # USD
    }
    async with httpx.AsyncClient(timeout=25.0) as client:
        r = await client.get(url, params=params)
        r.raise_for_status()
    root = ET.fromstring(r.text)
    rows: list[dict] = []
    for rec in root.findall(".//Record"):
        d = _parse_date_ru(rec.attrib.get("Date", ""))
        v = _parse_decimal((rec.findtext("Value") or "").strip())
        nominal = _parse_decimal((rec.findtext("Nominal") or "1").strip()) or 1.0
        if not d or v is None:
            continue
        rows.append({"date": d.isoformat(), "rate": round(v / nominal, 6)})
    rows.sort(key=lambda x: x["date"], reverse=True)
    return rows


def _parse_ozon_fx_tables_from_html(html: str) -> tuple[list[dict], list[dict]]:
    parser = _SimpleHtmlTableParser()
    parser.feed(html or "")

    def rows_to_rates(rows: list[list[str]]) -> list[dict]:
        out: list[dict] = []
        for row in rows:
            if len(row) < 2:
                continue
            d = _parse_date_ru(row[0])
            rate = _parse_decimal(row[1])
            if d and rate is not None:
                out.append({"date": d.isoformat(), "rate": rate})
        out.sort(key=lambda x: x["date"], reverse=True)
        return out

    sales_rows: list[dict] = []
    services_rows: list[dict] = []

    for table in parser.tables:
        rows = table.get("rows") or []
        if len(rows) < 2:
            continue
        flat_text = " ".join(" ".join(r).lower() for r in rows[:4])
        context = str(table.get("context") or "")
        marker = f"{context} {flat_text}"
        parsed = rows_to_rates(rows[1:] if rows and any("дата" in (c or "").lower() for c in rows[0]) else rows)
        if not parsed:
            continue
        if ("для продаж" in marker) or ("sales" in marker and not sales_rows):
            sales_rows = parsed
        elif ("для услуг" in marker) or ("services" in marker and not services_rows):
            services_rows = parsed

    # Fallback: if labels weren't detected, take first 2 date/rate tables.
    if not sales_rows or not services_rows:
        candidate_tables: list[list[dict]] = []
        for table in parser.tables:
            rows = table.get("rows") or []
            parsed = rows_to_rates(rows[1:] if rows and any("дата" in (c or "").lower() for c in rows[0]) else rows)
            if parsed:
                candidate_tables.append(parsed)
        if candidate_tables and not sales_rows:
            sales_rows = candidate_tables[0]
        if len(candidate_tables) > 1 and not services_rows:
            services_rows = candidate_tables[1]
    return sales_rows, services_rows


def _parse_ozon_fx_from_xlsx(content: bytes) -> tuple[list[dict], list[dict]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    def _sheet_kind(ws) -> str:
        title = str(ws.title or "").lower()
        if "прод" in title or "sale" in title:
            return "sales"
        if "услуг" in title or "service" in title:
            return "services"
        # inspect first rows text
        sample = []
        for row in ws.iter_rows(min_row=1, max_row=min(12, ws.max_row), values_only=True):
            sample.extend([str(v).strip().lower() for v in row if v is not None][:4])
        joined = " ".join(sample)
        if "для продаж" in joined:
            return "sales"
        if "для услуг" in joined:
            return "services"
        return ""

    def _extract_rows(ws) -> list[dict]:
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        date_col = None
        usd_col = None
        for i, row in enumerate(rows):
            cells = [str(v).strip().lower() if v is not None else "" for v in row]
            for idx, c in enumerate(cells):
                if c == "дата":
                    date_col = idx
            for idx, c in enumerate(cells):
                if c in {"$", "$ ", "usd"}:
                    usd_col = idx
                    break
            if date_col is not None and usd_col is not None:
                header_idx = i
                break
            date_col = None
            usd_col = None
        if header_idx is None or date_col is None or usd_col is None:
            return []

        out: list[dict] = []
        for row in rows[header_idx + 1 :]:
            if not row:
                continue
            d_raw = row[date_col] if date_col < len(row) else None
            u_raw = row[usd_col] if usd_col < len(row) else None
            if d_raw is None and u_raw is None:
                continue
            if isinstance(d_raw, datetime.datetime):
                d = d_raw.date()
            elif isinstance(d_raw, datetime.date):
                d = d_raw
            else:
                d = _parse_date_ru(d_raw)
            rate = _parse_decimal(u_raw)
            if not d or rate is None:
                continue
            out.append({"date": d.isoformat(), "rate": rate})
        out.sort(key=lambda x: x["date"], reverse=True)
        return out

    sales_rows: list[dict] = []
    services_rows: list[dict] = []
    fallback_tables: list[list[dict]] = []
    for ws in wb.worksheets:
        parsed = _extract_rows(ws)
        if not parsed:
            continue
        kind = _sheet_kind(ws)
        if kind == "sales" and not sales_rows:
            sales_rows = parsed
        elif kind == "services" and not services_rows:
            services_rows = parsed
        else:
            fallback_tables.append(parsed)

    if not sales_rows and fallback_tables:
        sales_rows = fallback_tables.pop(0)
    if not services_rows and fallback_tables:
        services_rows = fallback_tables.pop(0)
    return sales_rows, services_rows


async def _fetch_ozon_fx_rates(date_from: datetime.date, date_to: datetime.date) -> tuple[list[dict], list[dict]]:
    url = "https://docs.ozon.ru/global/accounting/receiving-payments/conversion-rate/?country=AE"
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": "https://docs.ozon.ru/",
    }
    async with httpx.AsyncClient(timeout=25.0, follow_redirects=True) as client:
        r = await client.get(url, headers=headers)
        r.raise_for_status()
    sales, services = _parse_ozon_fx_tables_from_html(r.text)
    frm = date_from.isoformat()
    to = date_to.isoformat()
    sales = [x for x in sales if frm <= x["date"] <= to]
    services = [x for x in services if frm <= x["date"] <= to]
    return sales, services


async def _fetch_ozon_fx_rates_playwright(date_from: datetime.date, date_to: datetime.date) -> tuple[list[dict], list[dict], dict]:
    if async_playwright is None:
        raise RuntimeError("Playwright не установлен в backend среде")
    url = "https://docs.ozon.ru/global/accounting/receiving-payments/conversion-rate/?country=AE"
    def _normalize_fx_rows(rows_raw: list[dict]) -> list[dict]:
        out: list[dict] = []
        for row in rows_raw or []:
            if not isinstance(row, dict):
                continue
            d = _parse_date_ru(row.get("date"))
            rate = _parse_decimal(row.get("usd"))
            if d and rate is not None:
                out.append({"date": d.isoformat(), "rate": rate})
        out.sort(key=lambda x: x["date"], reverse=True)
        return out

    debug: dict = {"engine": "playwright", "widget_found": False, "sales_rows_raw": 0, "services_rows_raw": 0}
    async with async_playwright() as p:  # type: ignore[misc]
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
            ),
            locale="ru-RU",
        )
        try:
            # Ускоряем загрузку: блокируем тяжелые ресурсы, нам нужен только HTML/JS с таблицами.
            await page.route(
                "**/*",
                lambda route: route.abort()
                if route.request.resource_type in {"image", "media", "font"}
                else route.continue_(),
            )
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            debug["final_url"] = page.url
            try:
                debug["title"] = await page.title()
            except Exception:
                debug["title"] = ""
            # Ждем именно виджет курсов валют.
            try:
                await page.wait_for_selector("#currency-vidget", timeout=12000)
                debug["widget_found"] = True
            except Exception:
                await page.wait_for_timeout(1500)
                try:
                    debug["widget_found"] = await page.locator("#currency-vidget").count() > 0
                except Exception:
                    debug["widget_found"] = False

            async def _extract_widget_rows() -> list[dict]:
                return await page.evaluate(
                    """
                    () => {
                      const root = document.querySelector('#currency-vidget');
                      if (!root) return [];
                      const tableBlock =
                        root.querySelector('[class*="_tableBlock"]') ||
                        root.querySelector('div[class*="tableBlock"]');
                      if (!tableBlock) return [];
                      const rowBlocks = Array.from(tableBlock.querySelectorAll('[class*="_tableMain"], div[class*="tableMain"]'));
                      const rows = [];
                      for (const rowEl of rowBlocks) {
                        const cells = Array.from(rowEl.querySelectorAll('[class*="_tableCell"], div[class*="tableCell"]'))
                          .map((el) => (el.textContent || '').trim());
                        if (!cells.length) continue;
                        const first = (cells[0] || '').toLowerCase();
                        if (first === 'дата') continue;
                        if (cells.length >= 2) {
                          rows.push({
                            date: cells[0] || '',
                            usd: cells[1] || '',
                            eur: cells[2] || '',
                            cny: cells[3] || '',
                          });
                        }
                      }
                      return rows;
                    }
                    """
                )

            async def _click_mode(mode_label: str) -> None:
                btn = page.get_by_role("button", name=mode_label)
                if await btn.count():
                    try:
                        await btn.first.click(timeout=5000)
                        await page.wait_for_timeout(500)
                    except Exception:
                        # Возможно кнопка уже активна/disabled
                        pass

            # Считываем обе вкладки виджета отдельно.
            await _click_mode("Для продаж")
            sales_raw = await _extract_widget_rows()
            debug["sales_rows_raw"] = len(sales_raw or [])
            await _click_mode("Для услуг")
            services_raw = await _extract_widget_rows()
            debug["services_rows_raw"] = len(services_raw or [])
        finally:
            await page.close()
            await browser.close()
    sales = _normalize_fx_rows(sales_raw if "sales_raw" in locals() else [])
    services = _normalize_fx_rows(services_raw if "services_raw" in locals() else [])
    # Fallback на HTML-парсер, если DOM extraction не сработал.
    if not sales and not services:
        html = ""
        try:
            # page уже закрыт, но если попадем сюда — значит extraction был пустым; повторим быстрым http fallback выше по стеку.
            pass
        except Exception:
            pass
        # На уровне Playwright HTML недоступен после закрытия страницы, поэтому просто вернем пусто.
        # Выше по стеку сработает кэш.
    frm = date_from.isoformat()
    to = date_to.isoformat()
    sales = [x for x in sales if frm <= x["date"] <= to]
    services = [x for x in services if frm <= x["date"] <= to]
    debug["sales_rows_filtered"] = len(sales)
    debug["services_rows_filtered"] = len(services)
    return sales, services, debug


def _catalog_marketplace_stores_context() -> list[dict]:
    integrations = load_integrations() or {}
    reg_map: dict[str, str] = {}
    source_items = load_sources() or []
    for source in source_items:
        source_id = str(source.get("id") or "").strip()
        if not source_id.startswith("pricing:"):
            continue
        table_name = str(get_registered_source_table(source_id) or "").strip()
        if table_name:
            reg_map[source_id] = table_name
    with _catalog_db_connect() as conn:
        if not reg_map:
            registry_rows = conn.execute(
                """
                SELECT source_id, table_name
                FROM source_tables_registry
                WHERE source_id LIKE 'pricing:%'
                """
            ).fetchall()
            for row in registry_rows:
                source_id = str(row["source_id"] or "").strip()
                table_name = str(row["table_name"] or "").strip()
                if source_id and table_name:
                    reg_map[source_id] = table_name
        db_store_rows = conn.execute(
            """
            SELECT store_uid, platform, store_id, store_name, currency_code, business_id, account_id, seller_id
            FROM stores
            ORDER BY platform, store_id
            """
        ).fetchall()
    stores: list[dict] = []

    ym = integrations.get("yandex_market") if isinstance(integrations.get("yandex_market"), dict) else {}
    for acc in (ym.get("accounts") or []):
        if not isinstance(acc, dict):
            continue
        business_id = str(acc.get("business_id") or "").strip()
        for shop in (acc.get("shops") or []):
            if not isinstance(shop, dict):
                continue
            campaign_id = str(shop.get("campaign_id") or "").strip()
            if not campaign_id:
                continue
            source_id = f"pricing:yandex_market:{campaign_id}"
            stores.append(
                {
                    "platform": "yandex_market",
                    "platform_label": "Яндекс.Маркет",
                    "store_id": campaign_id,
                    "store_uid": f"yandex_market:{campaign_id}",
                    "label": str(shop.get("campaign_name") or f"Магазин {campaign_id}"),
                    "store_name": str(shop.get("campaign_name") or f"Магазин {campaign_id}"),
                    "currency_code": str(shop.get("currency_code") or acc.get("currency_code") or "RUB").strip().upper() or "RUB",
                    "business_id": business_id,
                    "account_id": business_id,
                    "seller_id": "",
                    "source_id": source_id,
                    "table_name": reg_map.get(source_id, ""),
                }
            )

    existing_uids = {
        str(item.get("store_uid") or "").strip()
        for item in stores
        if str(item.get("store_uid") or "").strip()
    }
    for row in db_store_rows:
        platform = str(row["platform"] or "").strip().lower()
        store_uid = str(row["store_uid"] or "").strip()
        store_id = str(row["store_id"] or "").strip()
        if platform != "yandex_market" or not store_uid or not store_id or store_uid in existing_uids:
            continue
        source_id = f"pricing:yandex_market:{store_id}"
        stores.append(
            {
                "platform": "yandex_market",
                "platform_label": "Яндекс.Маркет",
                "store_id": store_id,
                "store_uid": store_uid,
                "label": str(row["store_name"] or f"Магазин {store_id}"),
                "store_name": str(row["store_name"] or f"Магазин {store_id}"),
                "currency_code": str(row["currency_code"] or "RUB").strip().upper() or "RUB",
                "business_id": str(row["business_id"] or "").strip(),
                "account_id": str(row["account_id"] or row["business_id"] or "").strip(),
                "seller_id": str(row["seller_id"] or "").strip(),
                "source_id": source_id,
                "table_name": reg_map.get(source_id, ""),
            }
        )

    oz = integrations.get("ozon") if isinstance(integrations.get("ozon"), dict) else {}
    for acc in (oz.get("accounts") or []):
        if not isinstance(acc, dict):
            continue
        client_id = str(acc.get("client_id") or "").strip()
        if not client_id:
            continue
        seller_id = str(acc.get("seller_id") or "").strip()
        seller_name = str(acc.get("seller_name") or f"Ozon кабинет {client_id}").strip()
        acc_stores = acc.get("stores") if isinstance(acc.get("stores"), list) else []
        own_store = next(
            (
                s for s in acc_stores
                if isinstance(s, dict) and str(s.get("store_id") or "").strip() == client_id
            ),
            None,
        )
        currency_code = str(
            (own_store or {}).get("currency_code")
            or acc.get("currency_code")
            or "RUB"
        ).strip().upper() or "RUB"
        source_id = f"pricing:ozon:{client_id}"
        stores.append(
            {
                "platform": "ozon",
                "platform_label": "Ozon",
                "store_id": client_id,
                "store_uid": f"ozon:{client_id}",
                "label": seller_name,
                "store_name": seller_name,
                "currency_code": currency_code,
                "business_id": "",
                "account_id": client_id,
                "seller_id": seller_id,
                "source_id": source_id,
                "table_name": reg_map.get(source_id, ""),
            }
        )
    stores.sort(key=lambda s: (s["platform_label"], str(s["label"]).lower(), str(s["store_id"])))
    return stores


def _catalog_external_tree_sources_context() -> list[dict]:
    rows = load_sources() or []
    out: list[dict] = []
    for src in rows:
        if not isinstance(src, dict):
            continue
        sid = str(src.get("id") or "").strip()
        if not sid:
            continue
        stype = str(src.get("type") or "").strip().lower()
        if stype not in {"gsheets", "yandex_tables", "external_system", "external"}:
            continue
        out.append({"id": sid, "type": stype, "label": str(src.get("name") or src.get("title") or sid)})
    out.sort(key=lambda x: (x["type"], x["label"].lower()))
    return out


def _upsert_google_account(name: str, service_account_json: str) -> dict:
    parsed, client_email, private_key_id = _validate_service_account_payload(service_account_json)

    data = load_integrations()
    google = data.get("google") if isinstance(data.get("google"), dict) else {}
    accounts = google.get("accounts") if isinstance(google.get("accounts"), list) else []
    now = _now_iso()

    existing = next(
        (a for a in accounts if isinstance(a, dict) and str(a.get("client_email") or "").strip() == client_email),
        None,
    )
    if existing:
        account_id = str(existing.get("id") or "") or uuid.uuid4().hex
        existing["id"] = account_id
        existing["name"] = name
        existing["service_account_json"] = service_account_json
        existing["service_account_b64"] = ""
        existing["private_key_id"] = private_key_id
        existing["updated_at"] = now
    else:
        account_id = uuid.uuid4().hex
        existing = {
            "id": account_id,
            "name": name,
            "client_email": client_email,
            "private_key_id": private_key_id,
            "service_account_json": service_account_json,
            "service_account_b64": "",
            "created_at": now,
            "updated_at": now,
        }
        accounts.append(existing)

    # Файл ключа на сервере (перезаписываем при повторной загрузке)
    key_file_path = GOOGLE_KEYS_DIR / f"{account_id}.json"
    key_file_path.write_text(json.dumps(parsed, ensure_ascii=False, indent=2), encoding="utf-8")
    existing["key_file_path"] = str(key_file_path.relative_to(BASE_DIR))

    google["accounts"] = accounts
    google["active_account_id"] = account_id
    # legacy compat
    google["service_account_json"] = service_account_json
    google["service_account_b64"] = ""
    data["google"] = google
    save_integrations(data)
    return {
        "id": account_id,
        "name": name,
        "client_email": client_email,
        "private_key_id": private_key_id,
    }


def _now_iso() -> str:
    return datetime.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _persist_shop_health(*, business_id: str, campaign_id: str, ok: bool, message: str = "") -> None:
    data = load_integrations()
    ym = data.get("yandex_market") or {}
    accounts = ym.get("accounts") if isinstance(ym.get("accounts"), list) else []
    bid = str(business_id or "").strip()
    cid = str(campaign_id or "").strip()
    now = _now_iso()

    changed = False
    for account in accounts:
        if not isinstance(account, dict):
            continue
        if str(account.get("business_id") or "").strip() != bid:
            continue
        shops = account.get("shops") if isinstance(account.get("shops"), list) else []
        for shop in shops:
            if not isinstance(shop, dict):
                continue
            if str(shop.get("campaign_id") or "").strip() != cid:
                continue
            shop["health_status"] = "ok" if ok else "error"
            shop["health_message"] = message or ""
            shop["health_checked_at"] = now
            changed = True
            break

    if changed:
        save_integrations(data)


def _persist_source_health(*, source_id: str, ok: bool, message: str = "") -> None:
    sources = load_sources()
    sid = str(source_id or "").strip()
    now = _now_iso()
    changed = False
    for src in sources:
        if str(src.get("id") or "").strip() != sid:
            continue
        src["health_status"] = "ok" if ok else "error"
        src["health_message"] = message or ""
        src["health_checked_at"] = now
        changed = True
        break
    if changed:
        save_sources(sources)


def _ym_headers(api_key: str) -> dict[str, str]:
    return {
        "Api-Key": api_key,
        "Accept": "application/json",
        "User-Agent": "pricing-analytics/1.0",
    }


def _find_yandex_shop_credentials(campaign_id: str) -> tuple[str, str, str] | None:
    cid = str(campaign_id or "").strip()
    if not cid:
        return None
    data = load_integrations()
    ym = data.get("yandex_market") if isinstance(data.get("yandex_market"), dict) else {}
    accounts = _normalize_ym_accounts(ym)
    for account in accounts:
        if not isinstance(account, dict):
            continue
        api_key = str(account.get("api_key") or "").strip()
        business_id = str(account.get("business_id") or "").strip()
        for shop in account.get("shops") or []:
            if str((shop or {}).get("campaign_id") or "").strip() == cid:
                return business_id, cid, api_key
    return None


def _find_ozon_account_credentials(client_id: str) -> tuple[str, str, str, str] | None:
    cid = str(client_id or "").strip()
    if not cid:
        return None
    data = load_integrations()
    oz = data.get("ozon") if isinstance(data.get("ozon"), dict) else {}
    accounts = _normalize_ozon_accounts(oz)
    for a in accounts:
        if str(a.get("client_id") or "").strip() != cid:
            continue
        api_key = str(a.get("api_key") or "").strip()
        seller_id = str(a.get("seller_id") or "").strip()
        seller_name = str(a.get("seller_name") or "").strip()
        return cid, api_key, seller_id, seller_name
    return None


def _extract_yandex_offer_items(payload: dict) -> tuple[list[dict], str]:
    root = payload if isinstance(payload, dict) else {}
    result = root.get("result") if isinstance(root.get("result"), dict) else root
    offers = result.get("offers") if isinstance(result, dict) and isinstance(result.get("offers"), list) else []
    paging = result.get("paging") if isinstance(result, dict) and isinstance(result.get("paging"), dict) else {}
    next_page_token = str(
        paging.get("nextPageToken")
        or paging.get("next_page_token")
        or result.get("nextPageToken")
        or root.get("nextPageToken")
        or ""
    ).strip()

    rows: list[dict] = []
    for raw in offers:
        if not isinstance(raw, dict):
            continue
        raw_status = str(
            raw.get("status")
            or raw.get("offerStatus")
            or raw.get("campaignOfferStatus")
            or ""
        ).upper()
        if raw.get("archived") is True or raw_status == "ARCHIVED":
            continue

        offer = raw.get("offer") if isinstance(raw.get("offer"), dict) else {}
        offer_content = raw.get("offerContent") if isinstance(raw.get("offerContent"), dict) else {}
        mapping = raw.get("mapping") if isinstance(raw.get("mapping"), dict) else {}
        category_obj = raw.get("category") if isinstance(raw.get("category"), dict) else {}

        sku = str(
            raw.get("offerId")
            or raw.get("shopSku")
            or offer.get("offerId")
            or offer.get("shopSku")
            or raw.get("sku")
            or mapping.get("shopSku")
            or ""
        ).strip()
        if not sku:
            continue

        name = str(
            raw.get("name")
            or offer_content.get("name")
            or offer.get("name")
            or raw.get("title")
            or mapping.get("name")
            or ""
        ).strip()

        category_name_raw = str(
            category_obj.get("name")
            or raw.get("categoryName")
            or raw.get("category")
            or ""
        ).strip()
        category_parent = str(
            category_obj.get("parentName")
            or category_obj.get("parent_name")
            or ""
        ).strip()
        category_path_raw = raw.get("categoryPath") or category_obj.get("path") or category_obj.get("fullPath") or ""
        path_parts: list[str] = []
        if isinstance(category_path_raw, list):
            path_parts = [str(x).strip() for x in category_path_raw if str(x).strip()]
        elif isinstance(category_path_raw, str) and category_path_raw.strip():
            sep = ">" if ">" in category_path_raw else ("/" if "/" in category_path_raw else None)
            if sep:
                path_parts = [p.strip() for p in category_path_raw.split(sep) if p.strip()]

        category = category_parent or (path_parts[-2] if len(path_parts) >= 2 else category_name_raw)
        subcategory = path_parts[-1] if len(path_parts) >= 2 else ""
        if not subcategory and category_parent and category_name_raw and category_name_raw != category_parent:
            subcategory = category_name_raw
        if not category:
            category = category_name_raw
        if not subcategory and len(path_parts) == 1:
            subcategory = path_parts[0]

        rows.append(
            {
                "sku": sku,
                "sku_original": sku,
                "name": name,
                "category": category,
                "subcategory": subcategory,
                "payload": raw,
            }
        )
    return rows, next_page_token


async def _fetch_yandex_campaign_offers(campaign_id: str, api_key: str) -> list[dict]:
    cid = str(campaign_id or "").strip()
    if not cid:
        return []

    page_token = ""
    collected: list[dict] = []
    seen_sku: set[str] = set()

    async with httpx.AsyncClient(timeout=45) as client:
        for _ in range(50):
            url = f"{YANDEX_BASE_URL}/campaigns/{cid}/offers"
            params: dict[str, str] = {"limit": "200"}
            if page_token:
                params["pageToken"] = page_token
            # limit/pageToken у метода передаются через query, фильтры — через body
            resp = await client.post(
                url,
                headers=_ym_headers(api_key),
                params=params,
                json={},
            )
            resp.raise_for_status()
            payload = resp.json()
            rows, next_page = _extract_yandex_offer_items(payload)
            for row in rows:
                sku = str(row.get("sku") or "").strip()
                if not sku or sku in seen_sku:
                    continue
                seen_sku.add(sku)
                collected.append(row)
            if not next_page or next_page == page_token:
                break
            page_token = next_page
    return collected


def _split_category_path(value: str) -> tuple[str, str]:
    raw = str(value or "").strip()
    if not raw:
        return "", ""
    separators = [" > ", " / ", ">", "/"]
    parts: list[str] = []
    for sep in separators:
        if sep in raw:
            parts = [p.strip() for p in raw.split(sep) if p.strip()]
            if parts:
                break
    if not parts:
        return raw, ""
    if len(parts) == 1:
        return parts[0], ""
    # Для настроек ценообразования:
    # category = верхний родительский уровень
    # subcategory = вся вложенная цепочка ниже
    return parts[0], " / ".join(parts[1:])


def _extract_yandex_offer_mappings(payload: dict) -> list[dict]:
    root = payload if isinstance(payload, dict) else {}
    result = root.get("result") if isinstance(root.get("result"), dict) else root
    rows = result.get("offerMappings")
    if not isinstance(rows, list):
        rows = result.get("offerMappingEntries")
    if not isinstance(rows, list):
        rows = []

    out: list[dict] = []
    for raw in rows:
        if not isinstance(raw, dict):
            continue
        offer = raw.get("offer") if isinstance(raw.get("offer"), dict) else {}
        mapping = raw.get("mapping") if isinstance(raw.get("mapping"), dict) else {}
        sku = str(
            offer.get("offerId")
            or offer.get("shopSku")
            or raw.get("offerId")
            or raw.get("shopSku")
            or ""
        ).strip()
        if not sku:
            continue
        name = str(
            offer.get("name")
            or mapping.get("marketSkuName")
            or raw.get("marketSkuName")
            or raw.get("name")
            or ""
        ).strip()
        category_raw = str(
            offer.get("category")
            or mapping.get("marketCategoryName")
            or raw.get("category")
            or raw.get("marketCategoryName")
            or ""
        ).strip()
        category, subcategory = _split_category_path(category_raw)
        market_category_id = str(
            mapping.get("marketCategoryId")
            or raw.get("marketCategoryId")
            or ""
        ).strip()

        def _find_dims_container() -> dict:
            candidates = [
                mapping.get("weightDimensions"),
                raw.get("weightDimensions"),
                offer.get("weightDimensions"),
                mapping.get("dimensions"),
                raw.get("dimensions"),
                offer.get("dimensions"),
            ]
            for c in candidates:
                if isinstance(c, dict):
                    return c
            # Некоторые ответы прячут габариты глубже.
            for container in (mapping, raw, offer):
                if not isinstance(container, dict):
                    continue
                for key, value in container.items():
                    lk = str(key).lower()
                    if ("weight" in lk or "dimension" in lk or "size" in lk) and isinstance(value, dict):
                        return value
            return {}

        dims = _find_dims_container()

        def _num_from(*values):
            for v in values:
                if v in (None, ""):
                    continue
                try:
                    return float(str(v).replace(",", "."))
                except Exception:
                    continue
            return None

        width_cm = _num_from(dims.get("width"), dims.get("widthCm"), dims.get("width_cm"))
        length_cm = _num_from(dims.get("length"), dims.get("lengthCm"), dims.get("length_cm"), dims.get("depth"))
        height_cm = _num_from(dims.get("height"), dims.get("heightCm"), dims.get("height_cm"))
        weight_kg = _num_from(
            dims.get("weight"),
            dims.get("weightKg"),
            dims.get("weight_kg"),
            dims.get("weightGross"),
        )
        out.append(
            {
                "sku": sku,
                "name": name,
                "category": category,
                "subcategory": subcategory,
                "market_category_id": market_category_id,
                "width_cm": width_cm,
                "length_cm": length_cm,
                "height_cm": height_cm,
                "weight_kg": weight_kg,
                "payload": raw,
            }
        )
    return out


def _walk_yandex_category_tree(node: dict, parents: list[str], out: dict[str, list[str]]) -> None:
    if not isinstance(node, dict):
        return
    name = str(node.get("name") or "").strip()
    current_path = [*parents, name] if name else list(parents)
    node_id = node.get("id")
    if node_id not in (None, ""):
        out[str(node_id)] = [p for p in current_path if p]
    children = node.get("children")
    if isinstance(children, list):
        for child in children:
            if isinstance(child, dict):
                _walk_yandex_category_tree(child, current_path, out)


def _normalize_yandex_category_path(path: list[str]) -> list[str]:
    technical_roots = {
        "все товары",
        "все категории",
        "товары",
        "каталог",
    }
    normalized = [str(p or "").strip() for p in path if str(p or "").strip()]
    while normalized and normalized[0].strip().lower() in technical_roots:
        normalized.pop(0)
    return normalized


async def _fetch_yandex_categories_tree_paths(api_key: str) -> dict[str, list[str]]:
    async with httpx.AsyncClient(timeout=45) as client:
        resp = await client.post(
            f"{YANDEX_BASE_URL}/categories/tree",
            headers=_ym_headers(api_key),
            json={"language": "RU"},
        )
        resp.raise_for_status()
        payload = resp.json()
    root = payload.get("result") if isinstance(payload, dict) and isinstance(payload.get("result"), dict) else {}
    out: dict[str, list[str]] = {}
    _walk_yandex_category_tree(root, [], out)
    return out


async def _fetch_yandex_offer_mappings_by_offer_ids(
    business_id: str,
    api_key: str,
    offer_ids: list[str],
) -> dict[str, dict]:
    bid = str(business_id or "").strip()
    if not bid or not offer_ids:
        return {}

    unique_offer_ids: list[str] = []
    seen: set[str] = set()
    for sku in offer_ids:
        val = str(sku or "").strip()
        if not val or val in seen:
            continue
        seen.add(val)
        unique_offer_ids.append(val)

    category_paths_by_id = await _fetch_yandex_categories_tree_paths(api_key)
    result: dict[str, dict] = {}
    async with httpx.AsyncClient(timeout=45) as client:
        for i in range(0, len(unique_offer_ids), 100):
            chunk = unique_offer_ids[i:i + 100]
            url = f"{YANDEX_BASE_URL}/businesses/{bid}/offer-mappings"
            resp = await client.post(
                url,
                headers=_ym_headers(api_key),
                json={"offerIds": chunk},
            )
            resp.raise_for_status()
            payload = resp.json()
            for row in _extract_yandex_offer_mappings(payload):
                sku = str(row.get("sku") or "").strip()
                market_category_id = str(row.get("market_category_id") or "").strip()
                if market_category_id and market_category_id in category_paths_by_id:
                    path = _normalize_yandex_category_path(category_paths_by_id[market_category_id])
                    if path:
                        row["category"] = path[0]
                        row["subcategory"] = " / ".join(path[1:]) if len(path) > 1 else ""
                if sku and sku not in result:
                    result[sku] = row
    return result


def _oz_headers(client_id: str, api_key: str) -> dict[str, str]:
    return {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "pricing-analytics/1.0",
    }


def _extract_ozon_product_list_items(payload: dict) -> tuple[list[dict], str]:
    root = payload if isinstance(payload, dict) else {}
    result = root.get("result") if isinstance(root.get("result"), dict) else {}
    items = result.get("items") if isinstance(result.get("items"), list) else []
    out: list[dict] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        product_id = str(item.get("product_id") or item.get("id") or "").strip()
        sku = str(item.get("offer_id") or item.get("offerId") or item.get("sku") or "").strip()
        if not product_id and not sku:
            continue
        out.append(
            {
                "product_id": product_id,
                "sku": sku,
                "name": str(item.get("name") or item.get("title") or "").strip(),
                "category_id": str(
                    item.get("category_id")
                    or item.get("description_category_id")
                    or item.get("descriptionCategoryId")
                    or ""
                ).strip(),
                "category": str(item.get("category_name") or item.get("category") or "").strip(),
                "payload": item,
            }
        )
    next_last_id = str(result.get("last_id") or result.get("lastId") or "").strip()
    return out, next_last_id


async def _fetch_ozon_product_list(client_id: str, api_key: str) -> list[dict]:
    collected: list[dict] = []
    seen_sku: set[str] = set()
    last_id = ""
    async with httpx.AsyncClient(timeout=45) as client:
        for _ in range(100):
            resp = await client.post(
                f"{OZON_BASE_URL}/v3/product/list",
                headers=_oz_headers(client_id, api_key),
                json={
                    "filter": {"visibility": "ALL"},
                    "last_id": last_id,
                    "limit": 100,
                },
            )
            resp.raise_for_status()
            payload = resp.json()
            rows, next_last_id = _extract_ozon_product_list_items(payload)
            for row in rows:
                sku = str(row.get("sku") or "").strip() or str(row.get("product_id") or "").strip()
                if not sku or sku in seen_sku:
                    continue
                seen_sku.add(sku)
                collected.append(row)
            if not next_last_id or next_last_id == last_id:
                break
            last_id = next_last_id
    return collected


def _extract_ozon_product_info_map(payload: dict) -> dict[str, dict]:
    root = payload if isinstance(payload, dict) else {}
    result = root.get("result")
    items = []
    if isinstance(result, dict) and isinstance(result.get("items"), list):
        items = result.get("items") or []
    elif isinstance(result, dict) and isinstance(result.get("products"), list):
        items = result.get("products") or []
    elif isinstance(result, list):
        items = result
    elif isinstance(root.get("items"), list):
        items = root.get("items") or []
    elif isinstance(root.get("products"), list):
        items = root.get("products") or []
    out: dict[str, dict] = {}

    def _collect_dict_nodes(obj: Any, depth: int = 0) -> list[dict]:
        if depth > 4:
            return []
        found: list[dict] = []
        if isinstance(obj, dict):
            found.append(obj)
            for v in obj.values():
                found.extend(_collect_dict_nodes(v, depth + 1))
        elif isinstance(obj, list):
            for v in obj:
                found.extend(_collect_dict_nodes(v, depth + 1))
        return found

    def _best_product_node(wrapper: dict) -> dict:
        candidates = _collect_dict_nodes(wrapper)
        best: tuple[int, dict] | None = None
        for node in candidates:
            score = 0
            if any(k in node for k in ("offer_id", "offerId", "sku")):
                score += 5
            if any(k in node for k in ("product_id", "id")):
                score += 4
            if any(k in node for k in ("description_category_id", "category_id", "descriptionCategoryId")):
                score += 4
            if any(k in node for k in ("name", "title")):
                score += 2
            if "commissions" in node:
                score += 3
            if best is None or score > best[0]:
                best = (score, node)
        return best[1] if best else wrapper

    for item in items:
        if not isinstance(item, dict):
            continue
        node = _best_product_node(item)
        product_id = str(
            node.get("id")
            or node.get("product_id")
            or item.get("id")
            or item.get("product_id")
            or ""
        ).strip()
        sku = str(
            node.get("offer_id")
            or node.get("offerId")
            or node.get("sku")
            or item.get("offer_id")
            or item.get("offerId")
            or item.get("sku")
            or ""
        ).strip()
        key = sku or product_id
        if not key:
            continue
        out[key] = {
            "product_id": product_id,
            "sku": sku,
            "name": str(
                node.get("name")
                or node.get("title")
                or item.get("name")
                or item.get("title")
                or ""
            ).strip(),
            "category_id": str(
                node.get("description_category_id")
                or node.get("category_id")
                or node.get("descriptionCategoryId")
                or item.get("description_category_id")
                or item.get("category_id")
                or item.get("descriptionCategoryId")
                or ""
            ).strip(),
            "category": str(
                node.get("description_category_name")
                or node.get("category_name")
                or node.get("category")
                or item.get("description_category_name")
                or item.get("category_name")
                or item.get("category")
                or ""
            ).strip(),
            "commission_percent": _extract_ozon_commission_percent(node) or _extract_ozon_commission_percent(item),
            "payload": item,
        }
    return out


def _extract_ozon_product_attributes_map(payload: dict) -> dict[str, dict]:
    root = payload if isinstance(payload, dict) else {}
    result = root.get("result")
    items: list[Any] = []
    if isinstance(result, list):
        items = result
    elif isinstance(result, dict):
        if isinstance(result.get("items"), list):
            items = result.get("items") or []
        elif isinstance(result.get("products"), list):
            items = result.get("products") or []
    elif isinstance(root.get("items"), list):
        items = root.get("items") or []

    out: dict[str, dict] = {}

    def _collect_dict_nodes(obj: Any, depth: int = 0) -> list[dict]:
        if depth > 4:
            return []
        found: list[dict] = []
        if isinstance(obj, dict):
            found.append(obj)
            for v in obj.values():
                found.extend(_collect_dict_nodes(v, depth + 1))
        elif isinstance(obj, list):
            for v in obj:
                found.extend(_collect_dict_nodes(v, depth + 1))
        return found

    def _pick_attr_node(wrapper: dict) -> dict:
        best: tuple[int, dict] | None = None
        for node in _collect_dict_nodes(wrapper):
            score = 0
            if any(k in node for k in ("offer_id", "offerId")):
                score += 4
            if any(k in node for k in ("product_id", "id")):
                score += 4
            if any(k in node for k in ("description_category_id", "descriptionCategoryId", "category_id")):
                score += 6
            if any(k in node for k in ("type_id", "description_type_id", "descriptionTypeId")):
                score += 6
            if any(k in node for k in ("name", "title")):
                score += 2
            if best is None or score > best[0]:
                best = (score, node)
        return best[1] if best else wrapper

    for item in items:
        if not isinstance(item, dict):
            continue
        node = _pick_attr_node(item)
        offer_id = str(
            node.get("offer_id")
            or node.get("offerId")
            or item.get("offer_id")
            or item.get("offerId")
            or ""
        ).strip()
        product_id = str(
            node.get("product_id")
            or node.get("id")
            or item.get("product_id")
            or item.get("id")
            or ""
        ).strip()
        key = offer_id or product_id
        if not key:
            continue
        out[key] = {
            "offer_id": offer_id,
            "product_id": product_id,
            "description_category_id": str(
                node.get("description_category_id")
                or node.get("descriptionCategoryId")
                or node.get("category_id")
                or item.get("description_category_id")
                or item.get("descriptionCategoryId")
                or item.get("category_id")
                or ""
            ).strip(),
            "type_id": str(
                node.get("type_id")
                or node.get("description_type_id")
                or node.get("descriptionTypeId")
                or item.get("type_id")
                or item.get("description_type_id")
                or item.get("descriptionTypeId")
                or ""
            ).strip(),
            "name": str(
                node.get("name")
                or node.get("title")
                or item.get("name")
                or item.get("title")
                or ""
            ).strip(),
            "payload": item,
        }
    return out


async def _fetch_ozon_product_attributes_map(
    client_id: str,
    api_key: str,
    offer_ids: list[str],
    product_ids: list[str] | None = None,
) -> dict[str, dict]:
    offers = [str(v or "").strip() for v in offer_ids if str(v or "").strip()]
    products = [str(v or "").strip() for v in (product_ids or []) if str(v or "").strip()]
    if not offers and not products:
        return {}

    out: dict[str, dict] = {}
    async with httpx.AsyncClient(timeout=45) as client:
        chunk_count = max((len(offers) + 99) // 100, (len(products) + 99) // 100, 1)
        for chunk_index in range(chunk_count):
            offer_chunk = offers[chunk_index * 100:(chunk_index + 1) * 100]
            product_chunk_raw = products[chunk_index * 100:(chunk_index + 1) * 100]
            product_chunk = [int(x) for x in product_chunk_raw if x.isdigit()]
            variants: list[tuple[str, dict[str, Any]]] = []
            # Основной сценарий: по артикулам продавца (offer_id = наш SKU).
            if offer_chunk:
                variants.append(
                    (
                        f"{OZON_BASE_URL}/v4/product/info/attributes",
                        {"filter": {"offer_id": offer_chunk}, "limit": len(offer_chunk)},
                    )
                )
                variants.append(
                    (
                        f"{OZON_BASE_URL}/v4/products/info/attributes",
                        {"filter": {"offer_id": offer_chunk}, "limit": len(offer_chunk)},
                    )
                )
            # Fallback: по product_id (Ozon ID товара)
            if product_chunk:
                variants.append(
                    (
                        f"{OZON_BASE_URL}/v4/product/info/attributes",
                        {"filter": {"product_id": product_chunk}, "limit": len(product_chunk)},
                    )
                )
                variants.append(
                    (
                        f"{OZON_BASE_URL}/v4/products/info/attributes",
                        {"filter": {"product_id": product_chunk}, "limit": len(product_chunk)},
                    )
                )

            parsed_any = False
            last_error: Exception | None = None
            for url, body in variants:
                try:
                    resp = await client.post(url, headers=_oz_headers(client_id, api_key), json=body)
                    resp.raise_for_status()
                    parsed = _extract_ozon_product_attributes_map(resp.json())
                    if not parsed:
                        body_json = resp.json()
                        root = body_json if isinstance(body_json, dict) else {}
                        result_obj = root.get("result")
                        result_type = type(result_obj).__name__
                        root_keys = list(root.keys())[:20] if isinstance(root, dict) else []
                        result_keys = list(result_obj.keys())[:20] if isinstance(result_obj, dict) else []
                        sample_items = None
                        if isinstance(result_obj, dict):
                            sample_items = result_obj.get("items") or result_obj.get("products")
                        elif isinstance(result_obj, list):
                            sample_items = result_obj
                        elif isinstance(root, dict):
                            sample_items = root.get("items") or root.get("products")
                        sample_keys = (
                            list(sample_items[0].keys())[:20]
                            if isinstance(sample_items, list) and sample_items and isinstance(sample_items[0], dict)
                            else []
                        )
                        last_error = RuntimeError(
                            "Пустой ответ ProductAttributesV4: "
                            f"{url}; root_keys={root_keys}; result_type={result_type}; "
                            f"result_keys={result_keys}; sample_item_keys={sample_keys}"
                        )
                        continue
                    out.update(parsed)
                    parsed_any = True
                    last_error = None
                    break
                except Exception as e:
                    last_error = e
                    continue
            if not parsed_any and last_error:
                raise RuntimeError(f"Не удалось загрузить атрибуты товаров Ozon (V4): {last_error}")
    return out


def _extract_ozon_commission_percent(item: dict) -> float | None:
    if not isinstance(item, dict):
        return None

    preferred_paths = [
        ("commissions", "sales_percent_fbo"),
        ("commissions", "sales_percent_fbs"),
        ("commissions", "sales_percent"),
        ("commissions", "fbo_sale_commission_percent"),
        ("commissions", "fbs_sale_commission_percent"),
    ]
    for path in preferred_paths:
        cur: Any = item
        ok = True
        for part in path:
            if isinstance(cur, dict) and part in cur:
                cur = cur[part]
            else:
                ok = False
                break
        if ok:
            try:
                return float(str(cur).replace(",", "."))
            except Exception:
                pass

    candidates: list[float] = []

    def walk(obj: Any, key_path: str = "") -> None:
        if isinstance(obj, dict):
            for k, v in obj.items():
                walk(v, f"{key_path}.{k}" if key_path else str(k))
            return
        if isinstance(obj, list):
            for i, v in enumerate(obj):
                walk(v, f"{key_path}[{i}]")
            return
        key = key_path.lower()
        if "comm" not in key:
            return
        if "percent" not in key and "rate" not in key:
            return
        try:
            candidates.append(float(str(obj).replace(",", ".")))
        except Exception:
            return

    walk(item)
    if not candidates:
        return None
    return max(candidates)


async def _fetch_ozon_product_info_map(
    client_id: str,
    api_key: str,
    product_ids: list[str],
    offer_ids: list[str] | None = None,
) -> dict[str, dict]:
    ids = [str(v or "").strip() for v in product_ids if str(v or "").strip()]
    offers = [str(v or "").strip() for v in (offer_ids or []) if str(v or "").strip()]
    if not ids and not offers:
        return {}
    out: dict[str, dict] = {}
    async with httpx.AsyncClient(timeout=45) as client:
        chunk_count = max((len(ids) + 99) // 100, (len(offers) + 99) // 100, 1)
        for chunk_index in range(chunk_count):
            chunk = ids[chunk_index * 100:(chunk_index + 1) * 100]
            offer_chunk = offers[chunk_index * 100:(chunk_index + 1) * 100]
            product_ids_int = [int(x) for x in chunk if x.isdigit()]
            request_variants: list[tuple[str, dict[str, Any]]] = []
            if offer_chunk:
                request_variants.append((f"{OZON_BASE_URL}/v3/product/info/list", {"offer_id": offer_chunk}))
            if product_ids_int:
                request_variants.append((f"{OZON_BASE_URL}/v3/product/info/list", {"product_id": product_ids_int}))

            if not request_variants:
                continue

            last_error: Exception | None = None
            parsed_any = False
            for url, body in request_variants:
                try:
                    resp = await client.post(url, headers=_oz_headers(client_id, api_key), json=body)
                    resp.raise_for_status()
                    parsed = _extract_ozon_product_info_map(resp.json())
                    if not parsed:
                        body_json = resp.json()
                        root = body_json if isinstance(body_json, dict) else {}
                        result_obj = root.get("result")
                        result_type = type(result_obj).__name__
                        root_keys = list(root.keys())[:20] if isinstance(root, dict) else []
                        result_keys = list(result_obj.keys())[:20] if isinstance(result_obj, dict) else []
                        sample_keys: list[str] = []
                        sample_items = None
                        if isinstance(result_obj, dict):
                            if isinstance(result_obj.get("items"), list):
                                sample_items = result_obj.get("items")
                            elif isinstance(result_obj.get("products"), list):
                                sample_items = result_obj.get("products")
                        elif isinstance(result_obj, list):
                            sample_items = result_obj
                        elif isinstance(root, dict) and isinstance(root.get("items"), list):
                            sample_items = root.get("items")
                        elif isinstance(root, dict) and isinstance(root.get("products"), list):
                            sample_items = root.get("products")
                        if isinstance(sample_items, list) and sample_items and isinstance(sample_items[0], dict):
                            sample_keys = list(sample_items[0].keys())[:20]
                        last_error = RuntimeError(
                            "Пустой ответ ProductInfoList: "
                            f"{url}; root_keys={root_keys}; result_type={result_type}; "
                            f"result_keys={result_keys}; sample_item_keys={sample_keys}"
                        )
                        continue
                    out.update(parsed)
                    parsed_any = True
                    last_error = None
                    break
                except Exception as e:
                    last_error = e
                    continue
            if not parsed_any and last_error:
                raise RuntimeError(f"Не удалось загрузить информацию о товарах Ozon: {last_error}")
    return out


def _walk_ozon_category_tree(
    node: dict,
    parents: list[str],
    out: dict[str, list[str]],
    nodes_out: list[dict[str, Any]] | None = None,
    parent_id: str = "",
) -> None:
    if not isinstance(node, dict):
        return

    # В Ozon в children могут лежать как категории, так и конечные type-узлы.
    direct_type_id = str(
        node.get("type_id")
        or node.get("description_type_id")
        or ""
    ).strip()
    direct_type_name = str(
        node.get("type_name")
        or node.get("description_type_name")
        or ""
    ).strip()
    if direct_type_id:
        type_path = [*parents, direct_type_name] if direct_type_name else list(parents)
        out[direct_type_id] = [p for p in type_path if p]
        if nodes_out is not None:
            nodes_out.append(
                {
                    "node_kind": "type",
                    "category_id": direct_type_id,
                    "type_id": direct_type_id,
                    "parent_category_id": str(parent_id or "").strip(),
                    "name": direct_type_name,
                    "type_name": direct_type_name,
                    "level": max(0, len(type_path) - 1),
                    "path": " / ".join([p for p in type_path if p]),
                }
            )
        # Для type-узлов дальше углубляться не нужно.
        return

    name = str(node.get("category_name") or node.get("name") or "").strip()
    cid = str(node.get("description_category_id") or node.get("category_id") or node.get("id") or "").strip()
    path = [*parents, name] if name else list(parents)
    if cid:
        out[cid] = [p for p in path if p]
        if nodes_out is not None:
            nodes_out.append(
                {
                    "node_kind": "category",
                    "category_id": cid,
                    "type_id": "",
                    "parent_category_id": str(parent_id or "").strip(),
                    "name": name,
                    "type_name": "",
                    "level": max(0, len(path) - 1),
                    "path": " / ".join([p for p in path if p]),
                }
            )
    children = node.get("children") or node.get("childrens")
    if isinstance(children, list):
        for child in children:
            if isinstance(child, dict):
                _walk_ozon_category_tree(child, path, out, nodes_out, cid)


async def _fetch_ozon_category_tree_snapshot(client_id: str, api_key: str) -> tuple[dict[str, list[str]], list[dict[str, Any]]]:
    async with httpx.AsyncClient(timeout=45) as client:
        resp = await client.post(
            f"{OZON_BASE_URL}/v1/description-category/tree",
            headers=_oz_headers(client_id, api_key),
            json={},
        )
        resp.raise_for_status()
        payload = resp.json()
    root = payload.get("result")
    nodes = root if isinstance(root, list) else ([root] if isinstance(root, dict) else [])
    out: dict[str, list[str]] = {}
    flat_nodes: list[dict[str, Any]] = []
    for node in nodes:
        if isinstance(node, dict):
            _walk_ozon_category_tree(node, [], out, flat_nodes)
    if not flat_nodes:
        root_type = type(root).__name__
        sample_keys: list[str] = []
        if isinstance(nodes, list) and nodes and isinstance(nodes[0], dict):
            sample_keys = list(nodes[0].keys())[:20]
        raise RuntimeError(
            "Пустой ответ дерева категорий Ozon: "
            f"result_type={root_type}; sample_node_keys={sample_keys}"
        )
    return out, flat_nodes


async def _fetch_ozon_seller_info(client_id: str, api_key: str) -> dict:
    endpoints = [
        ("POST", f"{OZON_BASE_URL}/v1/seller/info"),
        ("POST", f"{OZON_BASE_URL}/v2/seller/info"),
        ("GET", f"{OZON_BASE_URL}/v1/seller/info"),
    ]
    last_error: Exception | None = None
    async with httpx.AsyncClient(timeout=30) as client:
        for method, url in endpoints:
            try:
                if method == "POST":
                    resp = await client.post(url, headers=_oz_headers(client_id, api_key), json={})
                else:
                    resp = await client.get(url, headers=_oz_headers(client_id, api_key))
                resp.raise_for_status()
                payload = resp.json()
                result = payload.get("result") if isinstance(payload, dict) else None
                base = result if isinstance(result, dict) else (payload if isinstance(payload, dict) else {})
                seller_id = str(
                    base.get("seller_id")
                    or base.get("sellerId")
                    or base.get("id")
                    or ""
                ).strip()
                seller_name = str(
                    base.get("company_name")
                    or base.get("companyName")
                    or base.get("name")
                    or ""
                ).strip()
                if not seller_id:
                    seller_id = client_id
                if not seller_name:
                    seller_name = f"Ozon кабинет {seller_id}"
                return {"seller_id": seller_id, "seller_name": seller_name, "raw": payload}
            except Exception as e:
                last_error = e
                continue
    raise RuntimeError(f"Не удалось получить SellerInfo Ozon: {last_error}")


def _normalize_ozon_accounts(ozon: dict) -> list[dict]:
    accounts = ozon.get("accounts")
    out: list[dict] = []
    if isinstance(accounts, list):
        for a in accounts:
            if not isinstance(a, dict):
                continue
            client_id = str(a.get("client_id") or "").strip()
            api_key = str(a.get("api_key") or "").strip()
            seller_id = str(a.get("seller_id") or "").strip()
            seller_name = str(a.get("seller_name") or "").strip()
            stores = a.get("stores") if isinstance(a.get("stores"), list) else []
            norm_stores = []
            for s in stores:
                if not isinstance(s, dict):
                    continue
                sid = str(s.get("store_id") or "").strip()
                if not sid:
                    continue
                norm_stores.append(
                    {
                        "store_id": sid,
                        "store_name": str(s.get("store_name") or "").strip(),
                        "currency_code": str(s.get("currency_code") or "RUB").strip().upper() or "RUB",
                        "fulfillment_model": str(s.get("fulfillment_model") or "FBO").strip().upper() or "FBO",
                        "connected_at": s.get("connected_at"),
                        "health_status": s.get("health_status"),
                        "health_message": s.get("health_message"),
                        "health_checked_at": s.get("health_checked_at"),
                    }
                )
            if not norm_stores and seller_id:
                norm_stores = [
                    {
                        "store_id": seller_id,
                        "store_name": seller_name or f"Ozon {seller_id}",
                        "currency_code": str(a.get("currency_code") or "RUB").strip().upper() or "RUB",
                        "connected_at": a.get("connected_at"),
                        "health_status": a.get("health_status"),
                        "health_message": a.get("health_message"),
                        "health_checked_at": a.get("health_checked_at"),
                    }
                ]
            out.append(
                {
                    "client_id": client_id,
                    "api_key": api_key,
                    "seller_id": seller_id,
                    "seller_name": seller_name,
                    "currency_code": str(a.get("currency_code") or "RUB").strip().upper() or "RUB",
                    "fulfillment_model": str(a.get("fulfillment_model") or "FBO").strip().upper() or "FBO",
                    "connected_at": a.get("connected_at"),
                    "import_enabled": a.get("import_enabled"),
                    "export_enabled": a.get("export_enabled"),
                    "health_status": a.get("health_status"),
                    "health_message": a.get("health_message"),
                    "health_checked_at": a.get("health_checked_at"),
                    "stores": norm_stores,
                }
            )
        return out

    legacy_client = str(ozon.get("client_id") or "").strip()
    legacy_api = str(ozon.get("api_key") or "").strip()
    legacy_seller = str(ozon.get("seller_id") or "").strip()
    legacy_name = str(ozon.get("seller_name") or "").strip()
    if legacy_client or legacy_api or legacy_seller:
        out.append(
            {
                "client_id": legacy_client,
                "api_key": legacy_api,
                "seller_id": legacy_seller,
                "seller_name": legacy_name,
                "connected_at": ozon.get("connected_at"),
                "import_enabled": ozon.get("import_enabled"),
                "export_enabled": ozon.get("export_enabled"),
                "health_status": ozon.get("health_status"),
                "health_message": ozon.get("health_message"),
                "health_checked_at": ozon.get("health_checked_at"),
                "stores": [
                    {
                        "store_id": legacy_seller or legacy_client,
                        "store_name": legacy_name or (f"Ozon {legacy_seller}" if legacy_seller else "Ozon"),
                        "currency_code": str(ozon.get("currency_code") or "RUB").strip().upper() or "RUB",
                        "fulfillment_model": str(ozon.get("fulfillment_model") or "FBO").strip().upper() or "FBO",
                        "connected_at": ozon.get("connected_at"),
                        "health_status": ozon.get("health_status"),
                        "health_message": ozon.get("health_message"),
                        "health_checked_at": ozon.get("health_checked_at"),
                    }
                ],
            }
        )
    return out


def _normalize_ym_shops(ym: dict) -> list[dict]:
    shops = ym.get("shops")
    if isinstance(shops, list):
        normalized = []
        for s in shops:
            if not isinstance(s, dict):
                continue
            cid = str(s.get("campaign_id") or "").strip()
            if not cid:
                continue
            normalized.append(
                {
                    "campaign_id": cid,
                    "campaign_name": str(s.get("campaign_name") or "").strip(),
                    "business_id": str(s.get("business_id") or "").strip(),
                    "currency_code": str(s.get("currency_code") or "RUB").strip().upper() or "RUB",
                    "connected_at": s.get("connected_at"),
                    "import_enabled": s.get("import_enabled"),
                    "export_enabled": s.get("export_enabled"),
                    "health_status": s.get("health_status"),
                    "health_message": s.get("health_message"),
                    "health_checked_at": s.get("health_checked_at"),
                }
            )
        return normalized

    # legacy single-shop fallback
    legacy_campaign = str(ym.get("campaign_id") or "").strip()
    legacy_business = str(ym.get("business_id") or "").strip()
    if legacy_campaign:
        return [
            {
                "campaign_id": legacy_campaign,
                "campaign_name": str(ym.get("campaign_name") or "").strip(),
                "business_id": legacy_business,
                "currency_code": str(ym.get("currency_code") or "RUB").strip().upper() or "RUB",
                "fulfillment_model": str(ym.get("fulfillment_model") or "FBO").strip().upper() or "FBO",
                "connected_at": ym.get("connected_at"),
                "import_enabled": ym.get("import_enabled"),
                "export_enabled": ym.get("export_enabled"),
                "health_status": ym.get("health_status"),
                "health_message": ym.get("health_message"),
                "health_checked_at": ym.get("health_checked_at"),
            }
        ]
    return []


def _normalize_ym_accounts(ym: dict) -> list[dict]:
    accounts = ym.get("accounts")
    if isinstance(accounts, list):
        out = []
        for a in accounts:
            if not isinstance(a, dict):
                continue
            business_id = str(a.get("business_id") or "").strip()
            api_key = str(a.get("api_key") or "").strip()
            shops = a.get("shops")
            if not isinstance(shops, list):
                shops = []
            norm_shops = []
            for s in shops:
                if not isinstance(s, dict):
                    continue
                cid = str(s.get("campaign_id") or "").strip()
                if not cid:
                    continue
                norm_shops.append(
                    {
                        "campaign_id": cid,
                        "campaign_name": str(s.get("campaign_name") or "").strip(),
                        "business_id": str(s.get("business_id") or business_id).strip(),
                        "currency_code": str(s.get("currency_code") or "RUB").strip().upper() or "RUB",
                        "fulfillment_model": str(s.get("fulfillment_model") or "FBO").strip().upper() or "FBO",
                        "connected_at": s.get("connected_at"),
                        "import_enabled": s.get("import_enabled"),
                        "export_enabled": s.get("export_enabled"),
                        "health_status": s.get("health_status"),
                        "health_message": s.get("health_message"),
                        "health_checked_at": s.get("health_checked_at"),
                    }
                )
            out.append(
                {
                    "business_id": business_id,
                    "api_key": api_key,
                    "connected_at": a.get("connected_at"),
                    "import_enabled": a.get("import_enabled"),
                    "export_enabled": a.get("export_enabled"),
                    "health_status": a.get("health_status"),
                    "health_message": a.get("health_message"),
                    "health_checked_at": a.get("health_checked_at"),
                    "shops": norm_shops,
                }
            )
        return out

    # legacy fallback: single account with flat shops
    shops = _normalize_ym_shops(ym)
    legacy_business = str(ym.get("business_id") or "").strip()
    legacy_api = str(ym.get("api_key") or "").strip()
    if legacy_business or legacy_api or shops:
        return [
            {
                "business_id": legacy_business,
                "api_key": legacy_api,
                "connected_at": ym.get("connected_at"),
                "import_enabled": ym.get("import_enabled"),
                "export_enabled": ym.get("export_enabled"),
                "health_status": ym.get("health_status"),
                "health_message": ym.get("health_message"),
                "health_checked_at": ym.get("health_checked_at"),
                "shops": shops,
            }
        ]
    return []


def _extract_campaigns(payload: dict, business_id: str) -> list[dict]:
    root = payload or {}
    campaigns = root.get("campaigns") or ((root.get("result") or {}).get("campaigns")) or []
    business_id_str = str(business_id).strip()
    out_all: list[dict] = []
    out_filtered: list[dict] = []

    for c in campaigns:
        c_id = c.get("id")
        if c_id is None:
            continue

        c_business = c.get("business") or {}
        c_business_id = c.get("businessId") or c_business.get("id")

        c_name = c.get("name") or c.get("domain") or f"Кампания {c_id}"
        row = {
            "id": str(c_id),
            "name": str(c_name),
            "business_id": str(c_business_id or ""),
        }
        out_all.append(row)

        if not business_id_str or str(c_business_id) == business_id_str:
            out_filtered.append(row)

    # fallback: если пользователь случайно ввел campaignId вместо businessId
    if not out_filtered and business_id_str:
        out_filtered = [r for r in out_all if r["id"] == business_id_str]

    return out_filtered


def _effective_flow(
    *,
    global_flow: dict,
    platform_flow: dict,
    account: dict | None = None,
    shop: dict | None = None,
) -> dict[str, bool]:
    def _pick(key: str, default: bool) -> bool:
        if isinstance(shop, dict) and isinstance(shop.get(key), bool):
            return bool(shop.get(key))
        if isinstance(account, dict) and isinstance(account.get(key), bool):
            return bool(account.get(key))
        if isinstance(platform_flow.get(key), bool):
            return bool(platform_flow.get(key))
        if isinstance(global_flow.get(key), bool):
            return bool(global_flow.get(key))
        return default

    return {
        "import_enabled": _pick("import_enabled", True),
        "export_enabled": _pick("export_enabled", False),
    }


def _default_catalog_import_config() -> dict:
    return {
        "selected_sources": [],
        "master_source": "",
        "mapping_by_source": {},
        "updated_at": None,
    }


def _sanitize_catalog_mapping(mp: dict) -> dict[str, str]:
    return {
        "sku": str(mp.get("sku") or "").strip(),
        "name": str(mp.get("name") or "").strip(),
        "category": str(mp.get("category") or "").strip(),
        "subcategory": str(mp.get("subcategory") or "").strip(),
        "category_l1": str(mp.get("category_l1") or "").strip(),
        "category_l2": str(mp.get("category_l2") or "").strip(),
        "category_l3": str(mp.get("category_l3") or "").strip(),
        "category_l4": str(mp.get("category_l4") or "").strip(),
        "category_l5": str(mp.get("category_l5") or "").strip(),
    }


def _normalize_catalog_source_ids(raw: list[str]) -> list[str]:
    out: list[str] = []
    for item in raw:
        sid = str(item or "").strip()
        if not sid:
            continue
        if sid in out:
            continue
        out.append(sid)
    return out


def _build_catalog_import_options(data: dict) -> dict:
    ym = data.get("yandex_market") if isinstance(data.get("yandex_market"), dict) else {}
    ym_accounts = _normalize_ym_accounts(ym)

    sources = load_sources()
    gsheets_sources = []
    for src in sources:
        if str(src.get("type") or "").strip().lower() != "gsheets":
            continue
        sid = str(src.get("id") or "").strip()
        if not sid:
            continue
        mode_import = bool(src.get("mode_import", True))
        gsheets_sources.append(
            {
                "id": sid,
                "source_key": f"gsheets:{sid}",
                "title": str(src.get("title") or sid),
                "worksheet": str(src.get("worksheet") or ""),
                "mode_import": mode_import,
                "mapping_configured": _mapping_configured(src),
                "ready": bool(mode_import),
                "ready_reason": "" if mode_import else "Для Google Sheets включите режим импорта",
                "last_refreshed": src.get("last_refreshed"),
            }
        )

    source_options: list[dict] = []
    for acc in ym_accounts:
        bid = str(acc.get("business_id") or "").strip()
        shops = acc.get("shops") if isinstance(acc.get("shops"), list) else []
        if not bid or not shops:
            continue
        source_options.append(
            {
                "id": f"yandex_market:{bid}",
                "label": f"Яндекс.Маркет: Business ID {bid}",
                "kind": "account",
                "ready": True,
                "ready_reason": "",
            }
        )

    oz = data.get("ozon") if isinstance(data.get("ozon"), dict) else {}
    oz_accounts = _normalize_ozon_accounts(oz)
    for acc in oz_accounts:
        cid = str(acc.get("client_id") or "").strip()
        if not cid:
            continue
        sname = str(acc.get("seller_name") or "").strip()
        sid = str(acc.get("seller_id") or "").strip()
        label_tail = f"{sname} ({sid})" if sname or sid else f"Client ID {cid}"
        source_options.append(
            {
                "id": f"ozon:{cid}",
                "label": f"Ozon: {label_tail}",
                "kind": "account",
                "ready": True,
                "ready_reason": "",
            }
        )

    # В выбор источников каталога добавляем только готовые Google Sheets.
    # Если маппинг не настроен, источник остаётся в блоке "Google таблицы",
    # но не попадает в импорт каталога до завершения настройки.
    source_options.extend(
        [
            {
                "id": s["source_key"],
                "label": f"Google Sheets: {s['title']}",
                "kind": "gsheets",
                "ready": True,
                "ready_reason": "",
            }
            for s in gsheets_sources
            if s.get("ready")
        ]
    )

    return {
        "source_options": source_options,
        "gsheets_sources": gsheets_sources,
        "yandex_accounts": [
            {
                "business_id": str(a.get("business_id") or ""),
                "shops": [
                    {
                        "campaign_id": str(s.get("campaign_id") or ""),
                        "campaign_name": str(s.get("campaign_name") or ""),
                        "currency_code": str(s.get("currency_code") or "RUB"),
                    }
                    for s in (a.get("shops") or [])
                ],
            }
            for a in ym_accounts
        ],
    }


def _evaluate_catalog_import_config(cfg: dict, options: dict) -> tuple[bool, list[str]]:
    selected = _normalize_catalog_source_ids(cfg.get("selected_sources") or [])
    master = str(cfg.get("master_source") or "").strip()
    mapping_by_source = cfg.get("mapping_by_source") if isinstance(cfg.get("mapping_by_source"), dict) else {}
    opt_by_id = {
        str(o.get("id") or ""): o
        for o in (options.get("source_options") or [])
        if isinstance(o, dict) and str(o.get("id") or "").strip()
    }
    blockers: list[str] = []

    if not selected:
        blockers.append("Выберите минимум один источник для каталога")
    if selected and not master:
        blockers.append("Выберите источник дерева категорий и наименований")
    if master and master not in selected:
        blockers.append("Источник дерева должен входить в список выбранных источников")

    for sid in selected:
        opt = opt_by_id.get(sid)
        if not opt:
            blockers.append(f"Источник {sid} недоступен")
            continue
        if not bool(opt.get("ready")):
            reason = str(opt.get("ready_reason") or "Источник не готов")
            blockers.append(f"{opt.get('label')}: {reason}")
            continue
        if sid.startswith("gsheets:"):
            m = mapping_by_source.get(sid) if isinstance(mapping_by_source.get(sid), dict) else {}
            has_tree = bool(str(m.get("category") or "").strip() or str(m.get("category_l1") or "").strip())
            required = ("sku", "name")
            missing = [k for k in required if not str(m.get(k) or "").strip()]
            if not has_tree:
                missing.append("category|category_l1")
            if missing:
                blockers.append(f"{opt.get('label')}: не заполнен маппинг полей ({', '.join(missing)})")

    return len(blockers) == 0, blockers


def _get_catalog_import_config(data: dict) -> dict:
    imports_obj = data.get("imports") if isinstance(data.get("imports"), dict) else {}
    catalog_obj = imports_obj.get("catalog") if isinstance(imports_obj.get("catalog"), dict) else {}
    cfg = _default_catalog_import_config()
    raw_selected = catalog_obj.get("selected_sources")
    cfg["selected_sources"] = (
        _normalize_catalog_source_ids(raw_selected) if isinstance(raw_selected, list) else []
    )
    if cfg["selected_sources"]:
        ym = data.get("yandex_market") if isinstance(data.get("yandex_market"), dict) else {}
        ym_accounts = _normalize_ym_accounts(ym)
        first_bid = str((ym_accounts[0] or {}).get("business_id") or "").strip() if ym_accounts else ""
        oz = data.get("ozon") if isinstance(data.get("ozon"), dict) else {}
        oz_accounts = _normalize_ozon_accounts(oz)
        first_cid = str((oz_accounts[0] or {}).get("client_id") or "").strip() if oz_accounts else ""
        remapped: list[str] = []
        for sid in cfg["selected_sources"]:
            if sid == "yandex_market" and first_bid:
                remapped.append(f"yandex_market:{first_bid}")
            elif sid == "ozon" and first_cid:
                remapped.append(f"ozon:{first_cid}")
            else:
                remapped.append(sid)
        cfg["selected_sources"] = _normalize_catalog_source_ids(remapped)
    cfg["master_source"] = str(catalog_obj.get("master_source") or "").strip()
    raw_mapping = catalog_obj.get("mapping_by_source")
    if isinstance(raw_mapping, dict):
        sanitized: dict[str, dict[str, str]] = {}
        for sid, mp in raw_mapping.items():
            sid_norm = str(sid or "").strip()
            if not sid_norm or not isinstance(mp, dict):
                continue
            sanitized[sid_norm] = _sanitize_catalog_mapping(mp)
        cfg["mapping_by_source"] = sanitized
    if cfg["master_source"] == "yandex_market":
        ym = data.get("yandex_market") if isinstance(data.get("yandex_market"), dict) else {}
        ym_accounts = _normalize_ym_accounts(ym)
        bid = str((ym_accounts[0] or {}).get("business_id") or "").strip() if ym_accounts else ""
        if bid:
            cfg["master_source"] = f"yandex_market:{bid}"
    elif cfg["master_source"] == "ozon":
        oz = data.get("ozon") if isinstance(data.get("ozon"), dict) else {}
        oz_accounts = _normalize_ozon_accounts(oz)
        cid = str((oz_accounts[0] or {}).get("client_id") or "").strip() if oz_accounts else ""
        if cid:
            cfg["master_source"] = f"ozon:{cid}"
    cfg["updated_at"] = catalog_obj.get("updated_at")

    # Backward compatibility со старым single-provider форматом
    if not cfg["selected_sources"]:
        legacy_provider = str(catalog_obj.get("provider") or "").strip()
        legacy_sheet_id = str(catalog_obj.get("gsheets_source_id") or "").strip()
        if legacy_provider == "yandex_market":
            ym = data.get("yandex_market") if isinstance(data.get("yandex_market"), dict) else {}
            accounts = _normalize_ym_accounts(ym)
            bid = str((accounts[0] or {}).get("business_id") or "").strip() if accounts else ""
            if bid:
                key = f"yandex_market:{bid}"
                cfg["selected_sources"] = [key]
                cfg["master_source"] = key
        elif legacy_provider == "google_sheets" and legacy_sheet_id:
            key = f"gsheets:{legacy_sheet_id}"
            cfg["selected_sources"] = [key]
            cfg["master_source"] = key
        elif legacy_provider == "ozon":
            oz = data.get("ozon") if isinstance(data.get("ozon"), dict) else {}
            accounts = _normalize_ozon_accounts(oz)
            cid = str((accounts[0] or {}).get("client_id") or "").strip() if accounts else ""
            if cid:
                key = f"ozon:{cid}"
                cfg["selected_sources"] = [key]
                cfg["master_source"] = key

    return cfg


async def _fetch_campaigns_payload(api_key: str) -> dict:
    endpoints = [
        f"{YANDEX_BASE_URL}/campaigns",  # старый/совместимый путь
        "https://api.partner.market.yandex.ru/campaigns",  # путь из актуальной reference-страницы
    ]
    last_error: Exception | None = None

    async with httpx.AsyncClient(timeout=30) as client:
        for url in endpoints:
            try:
                resp = await client.get(url, headers=_ym_headers(api_key))
                resp.raise_for_status()
                return resp.json()
            except Exception as e:
                last_error = e
                continue

    raise RuntimeError(f"Не удалось получить campaigns: {last_error}")
