from __future__ import annotations

import datetime
import io
import logging
import re

import openpyxl
from fastapi import APIRouter, File, Form, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse

from backend.routers._shared import (
    _catalog_marketplace_stores_context,
    _fetch_cbr_usd_rates,
    _fetch_ozon_category_tree_snapshot,
    _fetch_ozon_product_attributes_map,
    _fetch_ozon_product_info_map,
    _fetch_ozon_product_list,
    _fetch_yandex_campaign_offers,
    _fetch_yandex_offer_mappings_by_offer_ids,
    _find_ozon_account_credentials,
    _find_yandex_shop_credentials,
    _parse_date_ymd,
    _read_source_rows,
    get_pricing_category_tree,
    get_pricing_store_settings,
    replace_category_tree_cache_nodes,
    get_category_tree_cache_paths,
    replace_pricing_category_tree,
    replace_source_rows,
    seed_pricing_category_settings_if_null,
    upsert_pricing_category_setting,
    upsert_pricing_store_settings,
    upsert_store,
    upsert_store_dataset,
    bulk_apply_pricing_defaults,
)
from backend.services.store_data_model import (
    clear_pricing_attractiveness_results_for_store,
    clear_pricing_boost_results_for_store,
    clear_pricing_promo_results_for_store,
    clear_pricing_price_results_for_store,
    clear_pricing_strategy_results_for_store,
    delete_dashboard_snapshots,
    get_dashboard_snapshot,
    get_monitoring_export_snapshot,
    get_pricing_logistics_product_settings_map,
    get_pricing_logistics_store_settings,
    upsert_dashboard_snapshot,
    upsert_refresh_job,
    upsert_pricing_logistics_product_settings_bulk,
    upsert_pricing_logistics_store_settings,
)
from backend.services.refresh_orchestrator_service import (
    configure_refresh_scheduler,
    get_refresh_monitoring_snapshot,
    start_refresh_all,
    start_refresh_job,
)
from backend.services.pricing_export_service import (
    export_strategy_outputs_for_all_stores,
    export_strategy_outputs_for_store,
)

router = APIRouter()
logger = logging.getLogger("uvicorn.error")

_PRICING_MONITORING_SNAPSHOT_NAME = "page_pricing_monitoring"
_PRICING_MONITORING_EXPORTS_SNAPSHOT_NAME = "page_pricing_monitoring_exports"
_PRICING_FX_RATES_SNAPSHOT_NAME = "page_pricing_fx_rates"

LOGISTICS_IMPORT_HEADERS = [
    "SKU",
    "Наименование товара",
    "Ширина, см",
    "Длина, см",
    "Высота, см",
    "Вес, кг",
    "Комментарии",
]


def _pricing_monitoring_snapshot_key() -> str:
    return f"{_PRICING_MONITORING_SNAPSHOT_NAME}:v2"


def _pricing_monitoring_exports_snapshot_key() -> str:
    return _PRICING_MONITORING_EXPORTS_SNAPSHOT_NAME


def _pricing_fx_rates_snapshot_key(*, period: str, date_from: str | None, date_to: str | None) -> str:
    return f"{period}|{str(date_from or '').strip()}|{str(date_to or '').strip()}"


def _invalidate_pricing_page_snapshots() -> None:
    delete_dashboard_snapshots(snapshot_name=_PRICING_MONITORING_SNAPSHOT_NAME)
    delete_dashboard_snapshots(snapshot_name=_PRICING_MONITORING_EXPORTS_SNAPSHOT_NAME)
    delete_dashboard_snapshots(snapshot_name=_PRICING_FX_RATES_SNAPSHOT_NAME)


def _build_pricing_monitoring_response() -> dict:
    response = get_refresh_monitoring_snapshot()
    upsert_dashboard_snapshot(
        snapshot_name=_PRICING_MONITORING_SNAPSHOT_NAME,
        cache_key=_pricing_monitoring_snapshot_key(),
        response=response,
    )
    return response


def _build_pricing_monitoring_exports_response() -> dict:
    response = get_monitoring_export_snapshot()
    upsert_dashboard_snapshot(
        snapshot_name=_PRICING_MONITORING_EXPORTS_SNAPSHOT_NAME,
        cache_key=_pricing_monitoring_exports_snapshot_key(),
        response=response,
    )
    return response


def _monitoring_snapshot_has_active_status(response: dict[str, object] | None) -> bool:
    payload = response if isinstance(response, dict) else {}
    run_all = payload.get("run_all")
    if isinstance(run_all, dict):
        status = str(run_all.get("last_status") or "").strip().lower()
        if status in {"running", "queued"}:
            return True
    rows = payload.get("rows")
    if not isinstance(rows, list):
        return False
    for row in rows:
        if not isinstance(row, dict):
            continue
        status = str(row.get("last_status") or "").strip().lower()
        if status in {"running", "queued"}:
            return True
    return False


def _invalidate_pricing_read_caches() -> None:
    from backend.routers.catalog import invalidate_catalog_cache
    from backend.services.pricing_attractiveness_service import invalidate_attractiveness_cache
    from backend.services.pricing_boost_service import invalidate_boost_cache
    from backend.services.pricing_prices_service import invalidate_prices_cache
    from backend.services.pricing_promos_service import invalidate_promos_cache
    from backend.services.pricing_strategy_service import invalidate_strategy_cache

    invalidate_catalog_cache()
    invalidate_prices_cache()
    invalidate_boost_cache()
    invalidate_attractiveness_cache()
    invalidate_promos_cache()
    invalidate_strategy_cache()


def _invalidate_pricing_materialized_results_for_store(*, store_uid: str) -> None:
    suid = str(store_uid or "").strip()
    if not suid:
        return
    clear_pricing_price_results_for_store(store_uid=suid)
    clear_pricing_boost_results_for_store(store_uid=suid)
    clear_pricing_attractiveness_results_for_store(store_uid=suid)
    clear_pricing_promo_results_for_store(store_uid=suid)
    clear_pricing_strategy_results_for_store(store_uid=suid)


@router.get("/api/pricing/settings/monitoring")
async def pricing_settings_monitoring():
    try:
        snapshot = get_dashboard_snapshot(
            snapshot_name=_PRICING_MONITORING_SNAPSHOT_NAME,
            cache_key=_pricing_monitoring_snapshot_key(),
        )
        if isinstance(snapshot, dict) and isinstance(snapshot.get("response"), dict):
            response = snapshot["response"]
            if not _monitoring_snapshot_has_active_status(response):
                return response
        return _build_pricing_monitoring_response()
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Не удалось получить мониторинг обновлений: {exc}"}, status_code=500)


@router.get("/api/pricing/settings/monitoring/exports")
async def pricing_settings_monitoring_exports():
    try:
        snapshot = get_dashboard_snapshot(
            snapshot_name=_PRICING_MONITORING_EXPORTS_SNAPSHOT_NAME,
            cache_key=_pricing_monitoring_exports_snapshot_key(),
        )
        if isinstance(snapshot, dict) and isinstance(snapshot.get("response"), dict):
            return snapshot["response"]
        return _build_pricing_monitoring_exports_response()
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Не удалось получить настройки экспорта: {exc}"}, status_code=500)


@router.post("/api/pricing/settings/monitoring/run-all")
async def pricing_settings_monitoring_run_all():
    try:
        result = start_refresh_all(trigger_source="manual")
        _invalidate_pricing_page_snapshots()
        _build_pricing_monitoring_response()
        _build_pricing_monitoring_exports_response()
        return result
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Не удалось запустить полный цикл обновлений: {exc}"}, status_code=500)


@router.post("/api/pricing/settings/monitoring/run-job")
async def pricing_settings_monitoring_run_job(payload: dict | None = None):
    body = payload if isinstance(payload, dict) else {}
    job_code = str(body.get("job_code") or "").strip()
    if not job_code:
        return JSONResponse({"ok": False, "message": "job_code обязателен"}, status_code=400)
    try:
        result = start_refresh_job(job_code, trigger_source="manual")
        _invalidate_pricing_page_snapshots()
        _build_pricing_monitoring_response()
        _build_pricing_monitoring_exports_response()
        return result
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Не удалось запустить задачу: {exc}"}, status_code=500)


@router.post("/api/pricing/settings/monitoring/export-config")
async def pricing_settings_monitoring_export_config(payload: dict | None = None):
    body = payload if isinstance(payload, dict) else {}
    store_uid = str(body.get("store_uid") or "").strip()
    export_kind = str(body.get("export_kind") or "").strip().lower()
    if not store_uid:
        return JSONResponse({"ok": False, "message": "store_uid обязателен"}, status_code=400)
    if export_kind not in {"prices", "ads"}:
        return JSONResponse({"ok": False, "message": "export_kind должен быть prices или ads"}, status_code=400)
    prefix = f"export_{export_kind}"
    try:
        upsert_pricing_store_settings(
            store_uid=store_uid,
            values={
                f"{prefix}_source_type": body.get("type"),
                f"{prefix}_source_id": body.get("sourceId"),
                f"{prefix}_source_name": body.get("sourceName"),
                f"{prefix}_sku_column": body.get("skuColumn"),
                f"{prefix}_value_column": body.get("valueColumn"),
            },
        )
        _invalidate_pricing_page_snapshots()
        _build_pricing_monitoring_exports_response()
        return {"ok": True}
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Не удалось сохранить настройки экспорта: {exc}"}, status_code=500)


@router.post("/api/pricing/settings/monitoring/export-run")
async def pricing_settings_monitoring_export_run(payload: dict | None = None):
    body = payload if isinstance(payload, dict) else {}
    store_uid = str(body.get("store_uid") or "").strip()
    try:
        if store_uid:
            result = await export_strategy_outputs_for_store(store_uid=store_uid)
        else:
            result = await export_strategy_outputs_for_all_stores()
        _invalidate_pricing_page_snapshots()
        _build_pricing_monitoring_exports_response()
        return {
            "ok": bool(result.get("ok")),
            "message": "" if bool(result.get("ok")) else "Экспорт завершился с ошибками",
            "result": result,
        }
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Не удалось выполнить экспорт: {exc}"}, status_code=500)


@router.post("/api/pricing/settings/monitoring/job")
async def pricing_settings_monitoring_upsert_job(payload: dict | None = None):
    body = payload if isinstance(payload, dict) else {}
    job_code = str(body.get("job_code") or "").strip()
    if not job_code:
        return JSONResponse({"ok": False, "message": "job_code обязателен"}, status_code=400)
    try:
        values = {
            "enabled": body.get("enabled"),
            "schedule_kind": body.get("schedule_kind"),
            "interval_minutes": body.get("interval_minutes"),
            "time_of_day": body.get("time_of_day"),
            "date_from": body.get("date_from"),
            "date_to": body.get("date_to"),
            "stores": body.get("stores"),
        }
        job = upsert_refresh_job(job_code=job_code, values=values)
        configure_refresh_scheduler()
        _invalidate_pricing_page_snapshots()
        _build_pricing_monitoring_response()
        return {"ok": True, "job": job}
    except Exception as exc:
        return JSONResponse({"ok": False, "message": f"Не удалось сохранить настройки обновления: {exc}"}, status_code=500)


def _normalize_header(v: str) -> str:
    return str(v or "").strip().lower()


def _num(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except Exception:
        try:
            return float(str(v).replace(",", "."))
        except Exception:
            return None


def _round(v: float | None, digits: int = 4):
    return None if v is None else round(float(v), digits)


def _upsert_marketplace_logistics_preserving_db(*, store_uid: str, rows: list[dict]) -> int:
    """
    Обновление логистических параметров из API маркетплейсов:
    если по SKU поле уже заполнено в БД, значение из API не перезаписывает его.
    """
    normalized_rows = []
    skus: list[str] = []
    for row in rows or []:
        sku = str((row or {}).get("sku") or "").strip()
        if not sku:
            continue
        skus.append(sku)
        normalized_rows.append(
            {
                "sku": sku,
                "width_cm": _num((row or {}).get("width_cm")),
                "length_cm": _num((row or {}).get("length_cm")),
                "height_cm": _num((row or {}).get("height_cm")),
                "weight_kg": _num((row or {}).get("weight_kg")),
            }
        )
    if not normalized_rows:
        return 0

    existing_map = get_pricing_logistics_product_settings_map(store_uid=store_uid, skus=skus)
    merged_rows: list[dict] = []
    for row in normalized_rows:
        sku = row["sku"]
        existing = existing_map.get(sku) or {}
        merged_rows.append(
            {
                "sku": sku,
                "width_cm": _num(existing.get("width_cm")) if _num(existing.get("width_cm")) is not None else row.get("width_cm"),
                "length_cm": _num(existing.get("length_cm")) if _num(existing.get("length_cm")) is not None else row.get("length_cm"),
                "height_cm": _num(existing.get("height_cm")) if _num(existing.get("height_cm")) is not None else row.get("height_cm"),
                "weight_kg": _num(existing.get("weight_kg")) if _num(existing.get("weight_kg")) is not None else row.get("weight_kg"),
            }
        )
    return upsert_pricing_logistics_product_settings_bulk(store_uid=store_uid, rows=merged_rows)


@router.get("/api/pricing/settings/market-items")
async def pricing_settings_market_items(platform: str, store_id: str):
    platform_id = str(platform or "").strip().lower()
    store = str(store_id or "").strip()
    if not platform_id or not store:
        return JSONResponse({"ok": False, "message": "platform и store_id обязательны"}, status_code=400)

    if platform_id == "ozon":
        creds = _find_ozon_account_credentials(store)
        if not creds:
            return JSONResponse({"ok": False, "message": f"Не найден кабинет Ozon для client_id={store}"}, status_code=404)
        client_id, api_key, seller_id, seller_name = creds
        if not api_key:
            return JSONResponse({"ok": False, "message": "У кабинета Ozon отсутствует API-ключ"}, status_code=400)
        try:
            category_paths_live, category_tree_nodes = await _fetch_ozon_category_tree_snapshot(client_id, api_key)
        except Exception as e:
            return JSONResponse({"ok": False, "message": f"Не удалось загрузить дерево категорий Ozon: {e}"}, status_code=502)
        try:
            base_rows = await _fetch_ozon_product_list(client_id, api_key)
        except Exception as e:
            return JSONResponse({"ok": False, "message": f"Не удалось загрузить товары Ozon: {e}"}, status_code=502)
        # Комиссии пробуем получить отдельно, но не ломаем построение категорий если не получилось.
        try:
            attrs_by_key = await _fetch_ozon_product_attributes_map(
                client_id,
                api_key,
                [str(r.get("sku") or "") for r in base_rows],
                [str(r.get("product_id") or "") for r in base_rows],
            )
        except Exception as e:
            return JSONResponse({"ok": False, "message": f"Не удалось загрузить атрибуты товаров Ozon: {e}"}, status_code=502)
        # Комиссии пробуем получить отдельно, но не ломаем построение категорий если не получилось.
        try:
            info_by_key = await _fetch_ozon_product_info_map(
                client_id,
                api_key,
                [str(r.get("product_id") or "") for r in base_rows],
                [str(r.get("sku") or "") for r in base_rows],
            )
        except Exception:
            info_by_key = {}
        category_paths: dict[str, list[str]] = {}
        if category_tree_nodes:
            try:
                replace_category_tree_cache_nodes(
                    platform="ozon",
                    account_id=client_id,
                    tree_code="description_categories",
                    nodes=category_tree_nodes,
                )
            except Exception as e:
                return JSONResponse(
                    {
                        "ok": False,
                        "message": f"Не удалось записать кэш дерева категорий Ozon: {e}",
                        "ozon_category_tree_nodes_parsed": len(category_tree_nodes),
                    },
                    status_code=502,
                )
        try:
            category_paths = get_category_tree_cache_paths(
                platform="ozon",
                account_id=client_id,
                tree_code="description_categories",
            )
        except Exception:
            category_paths = {}
        if not category_paths:
            category_paths = category_paths_live

        enriched_rows: list[dict] = []
        for row in base_rows:
            product_id = str(row.get("product_id") or "").strip()
            sku = str(row.get("sku") or "").strip()
            attrs = attrs_by_key.get(sku) or attrs_by_key.get(product_id) or {}
            info = info_by_key.get(sku) or info_by_key.get(product_id) or {}
            merged = dict(row)
            if str(attrs.get("name") or "").strip():
                merged["name"] = str(attrs.get("name") or "").strip()
            elif str(info.get("name") or "").strip():
                merged["name"] = str(info.get("name") or "").strip()
            category_id = str(attrs.get("description_category_id") or info.get("category_id") or row.get("category_id") or "").strip()
            type_id = str(attrs.get("type_id") or "").strip()
            merged["category_id"] = category_id
            merged["type_id"] = type_id
            # Предпочитаем путь типа товара (финальная подкатегория), затем путь категории.
            path = (category_paths.get(type_id) if type_id else None) or category_paths.get(category_id) or []
            if path:
                category = path[0]
                sub_parts = path[1:] if len(path) > 1 else []
                merged["category"] = category
                merged["subcategory"] = " / ".join([p for p in sub_parts if str(p).strip()])
            else:
                category_name = str(info.get("category") or row.get("category") or "").strip()
                if category_name:
                    merged["category"] = category_name
                    merged["subcategory"] = ""
            merged["payload"] = {
                "product_list": row.get("payload") if isinstance(row.get("payload"), dict) else {},
                "product_attrs_v4": attrs.get("payload") if isinstance(attrs.get("payload"), dict) else {},
                "product_info": info.get("payload") if isinstance(info.get("payload"), dict) else {},
            }
            if info.get("commission_percent") is not None:
                merged["commission_percent"] = info.get("commission_percent")
            enriched_rows.append(merged)

        source_id = f"pricing:ozon:{client_id}"
        try:
            saved_count = replace_source_rows(
                source_id,
                enriched_rows,
                source_type="pricing_ozon_products",
                title=f"Pricing Ozon {client_id}",
            )
        except Exception:
            saved_count = 0

        items = [
            {
                "sku": str(r.get("sku") or ""),
                "name": str(r.get("name") or ""),
                "category": str(r.get("category") or ""),
                "subcategory": str(r.get("subcategory") or ""),
                "commission_percent": (
                    None
                    if (r.get("commission_percent") is None or str(r.get("commission_percent")).strip() == "")
                    else r.get("commission_percent")
                ),
            }
            for r in enriched_rows
        ]

        grouped: dict[tuple[str, tuple[str, ...]], int] = {}
        grouped_commissions: dict[tuple[str, tuple[str, ...]], list[float]] = {}
        for item in items:
            category = str(item.get("category") or "").strip() or "Без категории"
            sub_levels = [p.strip() for p in str(item.get("subcategory") or "").split("/") if p.strip()][:5]
            key = (category, tuple(sub_levels))
            grouped[key] = grouped.get(key, 0) + 1
            try:
                cv_raw = item.get("commission_percent")
                if cv_raw not in (None, ""):
                    grouped_commissions.setdefault(key, []).append(float(str(cv_raw).replace(",", ".")))
            except Exception:
                pass

        try:
            store_uid = upsert_store(
                platform="ozon",
                store_id=client_id,
                store_name=seller_name or f"Ozon {client_id}",
                seller_id=seller_id,
                account_id=client_id,
            )
            dataset_key = upsert_store_dataset(
                store_uid=store_uid,
                store_id=client_id,
                task_code="pricing_categories",
                title=f"{client_id}: pricing_categories",
                status="ready",
                row_count=len(grouped),
                meta={
                    "platform": "ozon",
                    "client_id": client_id,
                    "seller_id": seller_id,
                    "source_id": source_id,
                    "raw_items_count": len(items),
                },
            )
            pricing_rows_saved = replace_pricing_category_tree(
                dataset_key=dataset_key,
                store_uid=store_uid,
                rows=[
                    {"category": cat, "subcategory_levels": list(levels), "items_count": cnt}
                    for (cat, levels), cnt in grouped.items()
                ],
            )
            api_seed_rows: list[dict] = []
            for (cat, levels), commissions in grouped_commissions.items():
                vals = sorted(v for v in commissions if isinstance(v, (int, float)))
                if not vals:
                    continue
                median = vals[len(vals) // 2]
                leaf_path = " / ".join([x for x in [cat, *list(levels)] if x])
                if not leaf_path:
                    continue
                api_seed_rows.append(
                    {
                        "leaf_path": leaf_path,
                        "commission_percent": round(float(median), 4),
                    }
                )
            if api_seed_rows:
                seed_pricing_category_settings_if_null(
                    dataset_key=dataset_key,
                    store_uid=store_uid,
                    rows=api_seed_rows,
                )
            # Слой логистики из Ozon: API заполняет только пустые поля, не перезаписывая ручные данные в БД.
            ozon_logistics_rows = []
            for r in enriched_rows:
                sku = str(r.get("sku") or "").strip()
                if not sku:
                    continue
                ozon_logistics_rows.append(
                    {
                        "sku": sku,
                        "width_cm": r.get("width_cm"),
                        "length_cm": r.get("length_cm"),
                        "height_cm": r.get("height_cm"),
                        "weight_kg": r.get("weight_kg"),
                    }
                )
            if ozon_logistics_rows:
                _upsert_marketplace_logistics_preserving_db(store_uid=store_uid, rows=ozon_logistics_rows)
        except Exception:
            store_uid = f"ozon:{client_id}"
            dataset_key = f"{client_id}:pricing_categories"
            pricing_rows_saved = 0

        type_nodes_count = 0
        try:
            type_nodes_count = sum(1 for n in category_tree_nodes if str((n or {}).get("node_kind") or "") == "type")
        except Exception:
            type_nodes_count = 0
        return {
            "ok": True,
            "platform": "ozon",
            "store_id": client_id,
            "seller_id": seller_id,
            "items": items,
            "total_count": len(items),
            "saved_count": saved_count,
            "store_uid": store_uid,
            "dataset_key": dataset_key,
            "pricing_category_rows_saved": pricing_rows_saved,
            "ozon_category_tree_nodes_saved": len(category_tree_nodes),
            "ozon_category_tree_type_nodes_saved": type_nodes_count,
        }

    if platform_id != "yandex_market":
        return {
            "ok": True,
            "platform": platform_id,
            "store_id": store,
            "items": [],
            "total_count": 0,
            "message": "Для этой площадки загрузка товаров в настройках ценообразования пока не реализована",
        }

    creds = _find_yandex_shop_credentials(store)
    if not creds:
        return JSONResponse({"ok": False, "message": f"Не найден кабинет Я.Маркета для campaign_id={store}"}, status_code=404)
    business_id, campaign_id, api_key = creds
    if not api_key:
        return JSONResponse({"ok": False, "message": "У кабинета Я.Маркета отсутствует API-ключ"}, status_code=400)

    try:
        rows = await _fetch_yandex_campaign_offers(campaign_id, api_key)
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"Не удалось загрузить товары Я.Маркета: {e}"}, status_code=502)

    try:
        mappings_by_sku = await _fetch_yandex_offer_mappings_by_offer_ids(
            business_id=business_id,
            api_key=api_key,
            offer_ids=[str(r.get("sku") or "") for r in rows],
        )
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"Не удалось загрузить маппинг товаров Я.Маркета: {e}"}, status_code=502)

    enriched_rows: list[dict] = []
    for row in rows:
        sku = str(row.get("sku") or "").strip()
        mapped = mappings_by_sku.get(sku) or {}
        merged = dict(row)
        if mapped:
            if str(mapped.get("name") or "").strip():
                merged["name"] = str(mapped.get("name") or "").strip()
            if str(mapped.get("category") or "").strip():
                merged["category"] = str(mapped.get("category") or "").strip()
            if str(mapped.get("subcategory") or "").strip():
                merged["subcategory"] = str(mapped.get("subcategory") or "").strip()
            payload_combined = {
                "campaign_offer": row.get("payload") if isinstance(row.get("payload"), dict) else {},
                "offer_mapping": mapped.get("payload") if isinstance(mapped.get("payload"), dict) else {},
            }
            merged["payload"] = payload_combined
            # Габариты/вес для логистики берём из актуального getOfferMappings (business-assortment).
            for dim_key in ("width_cm", "length_cm", "height_cm", "weight_kg"):
                if mapped.get(dim_key) is not None:
                    merged[dim_key] = mapped.get(dim_key)
        enriched_rows.append(merged)

    source_id = f"pricing:yandex_market:{campaign_id}"
    try:
        saved_count = replace_source_rows(
            source_id,
            enriched_rows,
            source_type="pricing_yandex_market_offers",
            title=f"Pricing Yandex Market {campaign_id}",
        )
    except Exception:
        saved_count = 0

    items = [
        {
            "sku": str(r.get("sku") or ""),
            "name": str(r.get("name") or ""),
            "category": str(r.get("category") or ""),
            "subcategory": str(r.get("subcategory") or ""),
        }
        for r in enriched_rows
    ]

    # Нормализованный SQL-слой для дальнейшей работы по магазину
    grouped: dict[tuple[str, tuple[str, ...]], int] = {}
    for item in items:
        category = str(item.get("category") or "").strip()
        sub_levels = [p.strip() for p in str(item.get("subcategory") or "").split("/") if p.strip()][:5]
        key = (category, tuple(sub_levels))
        grouped[key] = grouped.get(key, 0) + 1

    try:
        store_uid = upsert_store(
            platform="yandex_market",
            store_id=campaign_id,
            store_name=f"Я.Маркет {campaign_id}",
            business_id=business_id,
            account_id=business_id,
        )
        dataset_key = upsert_store_dataset(
            store_uid=store_uid,
            store_id=campaign_id,
            task_code="pricing_categories",
            title=f"{campaign_id}: pricing_categories",
            status="ready",
            row_count=len(grouped),
            meta={
                "platform": "yandex_market",
                "business_id": business_id,
                "campaign_id": campaign_id,
                "source_id": source_id,
                "raw_items_count": len(items),
            },
        )
        category_rows = [
            {
                "category": cat,
                "subcategory_levels": list(levels),
                "items_count": cnt,
            }
            for (cat, levels), cnt in grouped.items()
        ]
        pricing_rows_saved = replace_pricing_category_tree(
            dataset_key=dataset_key,
            store_uid=store_uid,
            rows=category_rows,
        )

        # Слой логистики по SKU: автозаполняем существующие параметры из Маркета.
        logistics_rows = []
        for r in enriched_rows:
            sku = str(r.get("sku") or "").strip()
            if not sku:
                continue
            logistics_rows.append(
                {
                    "sku": sku,
                    "width_cm": r.get("width_cm"),
                    "length_cm": r.get("length_cm"),
                    "height_cm": r.get("height_cm"),
                    "weight_kg": r.get("weight_kg"),
                }
            )
        if logistics_rows:
            _upsert_marketplace_logistics_preserving_db(store_uid=store_uid, rows=logistics_rows)
    except Exception:
        store_uid = f"yandex_market:{campaign_id}"
        dataset_key = f"{campaign_id}:pricing_categories"
        pricing_rows_saved = 0

    return {
        "ok": True,
        "platform": "yandex_market",
        "store_id": campaign_id,
        "business_id": business_id,
        "items": items,
        "total_count": len(items),
        "saved_count": saved_count,
        "store_uid": store_uid,
        "dataset_key": dataset_key,
        "pricing_category_rows_saved": pricing_rows_saved,
    }


@router.get("/api/pricing/settings/category-tree")
async def pricing_settings_category_tree(platform: str, store_id: str):
    platform_id = str(platform or "").strip().lower()
    sid = str(store_id or "").strip()
    if not platform_id or not sid:
        return JSONResponse({"ok": False, "message": "platform и store_id обязательны"}, status_code=400)
    store_uid = f"{platform_id}:{sid}"
    try:
        data = get_pricing_category_tree(store_uid=store_uid)
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"Не удалось загрузить дерево категорий: {e}"}, status_code=500)
    dataset_key = str(data.get("dataset_key") or "").strip()
    return {
        "ok": True,
        "store_uid": store_uid,
        "dataset_key": dataset_key,
        "rows": list(data.get("rows") or []),
        "total_count": len(list(data.get("rows") or [])),
        "store_settings": get_pricing_store_settings(store_uid=store_uid),
    }


@router.get("/api/pricing/settings/sales-plan")
async def pricing_settings_sales_plan():
    try:
        stores = _catalog_marketplace_stores_context()
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"Не удалось загрузить магазины: {e}"}, status_code=500)

    rows: list[dict] = []
    for store in stores:
        platform_id = str(store.get("platform") or "").strip().lower()
        store_id = str(store.get("store_id") or "").strip()
        if not platform_id or not store_id:
            continue
        store_uid = str(store.get("store_uid") or f"{platform_id}:{store_id}")
        settings = get_pricing_store_settings(store_uid=store_uid)
        rows.append(
            {
                "store_uid": store_uid,
                "platform": platform_id,
                "platform_label": str(store.get("platform_label") or platform_id),
                "store_id": store_id,
                "store_name": str(store.get("store_name") or store.get("label") or store_id),
                "currency_code": str(store.get("currency_code") or "RUB").strip().upper() or "RUB",
                "earning_mode": settings.get("earning_mode") or "profit",
                "strategy_mode": settings.get("strategy_mode") or "mix",
                "planned_revenue": settings.get("planned_revenue"),
                "target_drr_percent": settings.get("target_drr_percent"),
                "target_profit_rub": settings.get("target_profit_rub"),
                "target_profit_percent": settings.get("target_profit_percent"),
                "minimum_profit_percent": settings.get("minimum_profit_percent"),
                "target_margin_rub": settings.get("target_margin_rub"),
                "target_margin_percent": settings.get("target_margin_percent"),
                "updated_at": settings.get("updated_at"),
            }
        )
    rows.sort(key=lambda item: (str(item.get("platform_label") or ""), str(item.get("store_name") or "")))
    return {"ok": True, "rows": rows}


@router.post("/api/pricing/settings/category-settings")
async def pricing_settings_save_category_setting(payload: dict):
    platform_id = str(payload.get("platform") or "").strip().lower()
    store_id = str(payload.get("store_id") or "").strip()
    leaf_path = str(payload.get("leaf_path") or "").strip()
    values = payload.get("values")
    if not platform_id or not store_id or not leaf_path:
        return JSONResponse({"ok": False, "message": "platform, store_id, leaf_path обязательны"}, status_code=400)
    if not isinstance(values, dict):
        return JSONResponse({"ok": False, "message": "values должен быть объектом"}, status_code=400)

    store_uid = f"{platform_id}:{store_id}"
    tree = get_pricing_category_tree(store_uid=store_uid)
    dataset_key = str(tree.get("dataset_key") or "").strip()
    if not dataset_key:
        return JSONResponse({"ok": False, "message": "Не найден dataset pricing_categories для выбранного магазина"}, status_code=404)

    try:
        upsert_pricing_category_setting(
            dataset_key=dataset_key,
            store_uid=store_uid,
            leaf_path=leaf_path,
            values=values,
        )
        _invalidate_pricing_materialized_results_for_store(store_uid=store_uid)
        _invalidate_pricing_read_caches()
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"Не удалось сохранить значение: {e}"}, status_code=500)

    return {"ok": True, "dataset_key": dataset_key, "store_uid": store_uid, "leaf_path": leaf_path}


@router.post("/api/pricing/settings/store-settings")
async def pricing_settings_save_store_settings(payload: dict):
    platform_id = str(payload.get("platform") or "").strip().lower()
    store_id = str(payload.get("store_id") or "").strip()
    values = payload.get("values")
    if not platform_id or not store_id:
        return JSONResponse({"ok": False, "message": "platform и store_id обязательны"}, status_code=400)
    if not isinstance(values, dict):
        return JSONResponse({"ok": False, "message": "values должен быть объектом"}, status_code=400)
    store_uid = f"{platform_id}:{store_id}"
    try:
        # Автосохранение верхнего блока может сработать раньше, чем загрузка товаров создаст запись в stores.
        upsert_store(
            platform=platform_id,
            store_id=store_id,
            store_name=str(payload.get("store_name") or "").strip(),
        )
        settings = upsert_pricing_store_settings(store_uid=store_uid, values=values)
        _invalidate_pricing_materialized_results_for_store(store_uid=store_uid)
        _invalidate_pricing_read_caches()
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"Не удалось сохранить настройки магазина: {e}"}, status_code=500)
    return {"ok": True, "store_uid": store_uid, "settings": settings}


@router.get("/api/pricing/settings/logistics")
async def pricing_settings_logistics(
    platform: str,
    store_id: str,
    search: str = "",
    category_path: str = "",
    page: int = 1,
    page_size: int = 50,
):
    platform_id = str(platform or "").strip().lower()
    store = str(store_id or "").strip()
    if not platform_id or not store:
        return JSONResponse({"ok": False, "message": "platform и store_id обязательны"}, status_code=400)
    if platform_id not in {"yandex_market", "ozon"}:
        return JSONResponse({"ok": False, "message": "Поддерживаются только yandex_market и ozon"}, status_code=400)

    stores = _catalog_marketplace_stores_context()
    target_store = next((s for s in stores if s["platform"] == platform_id and s["store_id"] == store), None)
    if not target_store:
        return JSONResponse({"ok": False, "message": "Магазин не найден в источниках"}, status_code=404)

    store_uid = str(target_store.get("store_uid") or f"{platform_id}:{store}")
    table_name = str(target_store.get("table_name") or "").strip()
    if not table_name:
        return {
            "ok": True,
            "store_uid": store_uid,
            "store": target_store,
            "store_settings": get_pricing_logistics_store_settings(store_uid=store_uid),
            "rows": [],
            "total_count": 0,
            "page": max(1, int(page or 1)),
            "page_size": max(1, min(int(page_size or 50), 500)),
        }

    try:
        raw_rows = _read_source_rows(table_name)
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"Не удалось прочитать raw-товары магазина: {e}"}, status_code=500)

    q = str(search or "").strip().lower()
    category_path_raw = str(category_path or "").strip()
    selected_category_parts = [part.strip() for part in category_path_raw.split(" / ") if part.strip()]

    def _row_tree_path(row: dict) -> list[str]:
        category = str(row.get("category") or "").strip()
        subcategory_raw = str(row.get("subcategory") or "").strip()
        sub_parts = [part.strip() for part in re.split(r"\s*/\s*", subcategory_raw) if part.strip()]
        parts = [part for part in [category, *sub_parts] if part]
        return parts or ["Не определено"]

    def _build_tree(paths: list[list[str]]) -> list[dict]:
        root: dict[str, dict] = {}
        for path_parts in paths:
            level = root
            for part in path_parts:
                node = level.setdefault(part, {"name": part, "children": {}})
                level = node["children"]

        def _normalize(level: dict[str, dict]) -> list[dict]:
            return [
                {
                    "name": key,
                    "children": _normalize(node["children"]),
                }
                for key, node in sorted(level.items(), key=lambda item: item[0].lower())
            ]

        return _normalize(root)

    searched_rows: list[dict] = []
    for r in raw_rows:
        sku = str(r.get("sku") or "").strip()
        if not sku:
            continue
        name = str(r.get("name") or "").strip()
        if q and q not in f"{sku} {name}".lower():
            continue
        searched_rows.append(r)

    filtered: list[dict] = []
    for r in searched_rows:
        tree_path_parts = _row_tree_path(r)
        if selected_category_parts and tree_path_parts[: len(selected_category_parts)] != selected_category_parts:
            continue
        filtered.append(r)
    filtered.sort(key=lambda r: str(r.get("sku") or ""))
    tree_roots = _build_tree([_row_tree_path(r) for r in searched_rows])

    page_size_n = max(1, min(int(page_size or 50), 500))
    page_n = max(1, int(page or 1))
    start = (page_n - 1) * page_size_n
    paged = filtered[start:start + page_size_n]

    sku_list = [str(r.get("sku") or "").strip() for r in paged if str(r.get("sku") or "").strip()]
    dim_map = get_pricing_logistics_product_settings_map(store_uid=store_uid, skus=sku_list)
    store_settings = get_pricing_logistics_store_settings(store_uid=store_uid)
    all_store_maps: dict[str, dict[str, dict]] = {}
    store_labels: dict[str, str] = {}
    for s in stores:
        suid = str(s.get("store_uid") or f"{s.get('platform')}:{s.get('store_id')}")
        store_labels[suid] = str(s.get("name") or s.get("store_id") or suid)
        try:
            all_store_maps[suid] = get_pricing_logistics_product_settings_map(store_uid=suid, skus=sku_list)
        except Exception:
            all_store_maps[suid] = {}

    divisor = 1000.0 if platform_id == "yandex_market" else 5000.0

    handling_mode = str(store_settings.get("handling_mode") or "fixed").strip().lower()
    handling_fixed = _num(store_settings.get("handling_fixed_amount"))
    handling_percent = _num(store_settings.get("handling_percent"))
    handling_min = _num(store_settings.get("handling_min_amount"))
    handling_max = _num(store_settings.get("handling_max_amount"))
    delivery_per_kg = _num(store_settings.get("delivery_cost_per_kg"))
    return_processing = _num(store_settings.get("return_processing_cost"))
    disposal_cost = _num(store_settings.get("disposal_cost"))

    def _handling_label() -> str:
        if handling_mode == "percent":
            p = f"{handling_percent:g}%" if handling_percent is not None else "—%"
            lo = f"{handling_min:g}" if handling_min is not None else "—"
            hi = f"{handling_max:g}" if handling_max is not None else "—"
            return f"{p} (мин {lo}, макс {hi})"
        return f"{handling_fixed:g}" if handling_fixed is not None else "—"

    def _handling_calc_value() -> float | None:
        # Для расчёта по Я.Маркет:
        # fixed -> фиксированная сумма
        # percent -> берём минимальный порог как рабочую базу до интеграции цены товара
        if handling_mode == "percent":
            return handling_min
        return handling_fixed

    rows_out: list[dict] = []
    for r in paged:
        sku = str(r.get("sku") or "").strip()
        name = str(r.get("name") or "").strip()
        dims_current = dim_map.get(sku) or {}
        inherited_fields: set[str] = set()
        comments: list[str] = []
        field_values_by_store: dict[str, dict[str, float]] = {}

        for field in ("width_cm", "length_cm", "height_cm", "weight_kg"):
            vals: dict[str, float] = {}
            for suid, m in all_store_maps.items():
                val = _num((m.get(sku) or {}).get(field))
                if val is not None:
                    vals[suid] = val
            field_values_by_store[field] = vals
            unique_vals = sorted({round(v, 6) for v in vals.values()})
            if len(unique_vals) > 1:
                label = {
                    "width_cm": "Ширина",
                    "length_cm": "Длина",
                    "height_cm": "Высота",
                    "weight_kg": "Вес",
                }.get(field, field)
                details = "; ".join(
                    f"{store_labels.get(suid, suid)}={vals[suid]:g}"
                    for suid in sorted(vals.keys(), key=lambda x: store_labels.get(x, x).lower())
                )
                comments.append(f"{label}: {details}")

        def _pick_dim(field: str) -> float | None:
            v_cur = _num(dims_current.get(field))
            if v_cur is not None:
                return v_cur
            vals = field_values_by_store.get(field) or {}
            if not vals:
                return None
            inherited_fields.add(field)
            first_uid = next(iter(vals.keys()))
            return vals.get(first_uid)

        width = _pick_dim("width_cm")
        length = _pick_dim("length_cm")
        height = _pick_dim("height_cm")
        weight = _pick_dim("weight_kg")
        volumetric = None
        if width is not None and length is not None and height is not None:
            volumetric = (width * length * height) / divisor
        max_weight = None
        if weight is not None and volumetric is not None:
            max_weight = max(weight, volumetric)
        elif weight is not None:
            max_weight = weight
        elif volumetric is not None:
            max_weight = volumetric
        delivery_to_client = None
        logistics_total = None
        if platform_id == "yandex_market":
            h_val = _handling_calc_value()
            if max_weight is not None and delivery_per_kg is not None:
                base_delivery = (delivery_per_kg * max_weight)
                if h_val is not None:
                    base_delivery += h_val
                delivery_to_client = base_delivery
            elif h_val is not None:
                delivery_to_client = h_val
            if delivery_to_client is not None:
                logistics_total = delivery_to_client + (return_processing or 0.0) + (disposal_cost or 0.0)
        rows_out.append(
            {
                "sku": sku,
                "name": name,
                "tree_path": _row_tree_path(r),
                "logistics_cost_display": _round(logistics_total, 4),
                "width_cm": _round(width, 3),
                "length_cm": _round(length, 3),
                "height_cm": _round(height, 3),
                "weight_kg": _round(weight, 3),
                "dimensions_inherited": bool(inherited_fields),
                "volumetric_weight_kg": _round(volumetric, 3),
                "max_weight_kg": _round(max_weight, 3),
                "cost_per_kg": _round(delivery_per_kg, 4),
                "handling_cost_display": _handling_label(),
                "delivery_to_client_cost": _round(delivery_to_client, 4),
                "return_processing_cost": _round(return_processing, 4),
                "disposal_cost": _round(disposal_cost, 4),
                "comments": "\n".join(comments),
                "updated_at": str(r.get("updated_at") or ""),
            }
        )

    return {
        "ok": True,
        "store_uid": store_uid,
        "store": target_store,
        "store_settings": store_settings,
        "tree_roots": tree_roots,
        "rows": rows_out,
        "total_count": len(filtered),
        "page": page_n,
        "page_size": page_size_n,
        "page_size_options": [25, 50, 100, 200, 500],
    }


@router.post("/api/pricing/settings/logistics/store-settings")
async def pricing_settings_save_logistics_store_settings(payload: dict):
    platform_id = str(payload.get("platform") or "").strip().lower()
    store_id = str(payload.get("store_id") or "").strip()
    values = payload.get("values")
    if not platform_id or not store_id:
        return JSONResponse({"ok": False, "message": "platform и store_id обязательны"}, status_code=400)
    if not isinstance(values, dict):
        return JSONResponse({"ok": False, "message": "values должен быть объектом"}, status_code=400)
    store_uid = f"{platform_id}:{store_id}"
    try:
        upsert_store(
            platform=platform_id,
            store_id=store_id,
            store_name=str(payload.get("store_name") or "").strip(),
        )
        settings = upsert_pricing_logistics_store_settings(store_uid=store_uid, values=values)
        _invalidate_pricing_materialized_results_for_store(store_uid=store_uid)
        _invalidate_pricing_read_caches()
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"Не удалось сохранить логистику магазина: {e}"}, status_code=500)
    return {"ok": True, "store_uid": store_uid, "settings": settings}


@router.post("/api/pricing/settings/logistics/product-settings")
async def pricing_settings_save_logistics_product_settings(payload: dict):
    platform_id = str(payload.get("platform") or "").strip().lower()
    store_id = str(payload.get("store_id") or "").strip()
    sku = str(payload.get("sku") or "").strip()
    values = payload.get("values")
    if not platform_id or not store_id or not sku:
        return JSONResponse({"ok": False, "message": "platform, store_id и sku обязательны"}, status_code=400)
    if not isinstance(values, dict):
        return JSONResponse({"ok": False, "message": "values должен быть объектом"}, status_code=400)
    allowed = {"width_cm", "length_cm", "height_cm", "weight_kg"}
    filtered = {k: values.get(k) for k in allowed if k in values}
    if not filtered:
        return JSONResponse({"ok": False, "message": "Нет допустимых полей для сохранения"}, status_code=400)
    store_uid = f"{platform_id}:{store_id}"
    try:
        upsert_store(
            platform=platform_id,
            store_id=store_id,
            store_name=str(payload.get("store_name") or "").strip(),
        )
        upsert_pricing_logistics_product_settings_bulk(
            store_uid=store_uid,
            rows=[{"sku": sku, **filtered}],
        )
        _invalidate_pricing_materialized_results_for_store(store_uid=store_uid)
        _invalidate_pricing_read_caches()
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"Не удалось сохранить параметры логистики SKU: {e}"}, status_code=500)
    return {"ok": True, "store_uid": store_uid, "sku": sku}


@router.get("/api/pricing/settings/logistics/import-template")
async def pricing_settings_logistics_import_template(platform: str, store_id: str):
    platform_id = str(platform or "").strip().lower()
    store = str(store_id or "").strip()
    if not platform_id or not store:
        return JSONResponse({"ok": False, "message": "platform и store_id обязательны"}, status_code=400)
    if platform_id not in {"yandex_market", "ozon"}:
        return JSONResponse({"ok": False, "message": "Поддерживаются только yandex_market и ozon"}, status_code=400)

    # Вытягиваем все строки постранично (эндпоинт логистики ограничивает размер страницы).
    all_rows: list[dict] = []
    page_n = 1
    page_size_n = 500
    total_count = None
    while True:
        data = await pricing_settings_logistics(
            platform=platform_id,
            store_id=store,
            search="",
            page=page_n,
            page_size=page_size_n,
        )
        if isinstance(data, JSONResponse):
            return data
        batch = list((data or {}).get("rows") or [])
        if total_count is None:
            try:
                total_count = int((data or {}).get("total_count") or 0)
            except Exception:
                total_count = 0
        if not batch:
            break
        all_rows.extend(batch)
        if total_count and len(all_rows) >= total_count:
            break
        page_n += 1
    rows = all_rows

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Логистика"
    ws.append(LOGISTICS_IMPORT_HEADERS)
    for row in rows:
        ws.append(
            [
                str(row.get("sku") or ""),
                str(row.get("name") or ""),
                row.get("width_cm"),
                row.get("length_cm"),
                row.get("height_cm"),
                row.get("weight_kg"),
                str(row.get("comments") or ""),
            ]
        )
    for col in ("A", "B", "C", "D", "E", "F", "G"):
        if col == "B":
            ws.column_dimensions[col].width = 28
        elif col == "G":
            ws.column_dimensions[col].width = 56
        else:
            ws.column_dimensions[col].width = 18

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"logistics_template_{platform_id}_{store}_{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/api/pricing/settings/logistics/import")
async def pricing_settings_logistics_import(
    platform: str = Form(...),
    store_id: str = Form(...),
    apply_scope: str = Form("store"),
    file: UploadFile = File(...),
):
    platform_id = str(platform or "").strip().lower()
    store = str(store_id or "").strip()
    scope = str(apply_scope or "store").strip().lower()
    if not platform_id or not store:
        return JSONResponse({"ok": False, "message": "platform и store_id обязательны"}, status_code=400)
    if platform_id not in {"yandex_market", "ozon"}:
        return JSONResponse({"ok": False, "message": "Поддерживаются только yandex_market и ozon"}, status_code=400)
    if scope not in {"store", "all"}:
        return JSONResponse({"ok": False, "message": "apply_scope должен быть store или all"}, status_code=400)

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"Не удалось прочитать Excel: {e}"}, status_code=400)

    ws = wb.active
    header = [str(c.value or "").strip() for c in ws[1][: len(LOGISTICS_IMPORT_HEADERS)]]
    if [_normalize_header(h) for h in header] != [_normalize_header(h) for h in LOGISTICS_IMPORT_HEADERS]:
        return JSONResponse(
            {
                "ok": False,
                "message": (
                    "Неверный порядок столбцов. Ожидается: "
                    + ", ".join(LOGISTICS_IMPORT_HEADERS)
                ),
            },
            status_code=400,
        )

    rows_to_apply: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        sku = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if not sku:
            continue
        width = _num(ws.cell(row=row_idx, column=3).value)
        length = _num(ws.cell(row=row_idx, column=4).value)
        height = _num(ws.cell(row=row_idx, column=5).value)
        weight = _num(ws.cell(row=row_idx, column=6).value)
        rows_to_apply.append(
            {
                "sku": sku,
                "width_cm": width,
                "length_cm": length,
                "height_cm": height,
                "weight_kg": weight,
            }
        )

    if not rows_to_apply:
        return JSONResponse({"ok": False, "message": "В файле нет строк для импорта"}, status_code=400)

    stores_ctx = _catalog_marketplace_stores_context()
    if scope == "all":
        target_uids = [
            str(s.get("store_uid") or f"{s.get('platform')}:{s.get('store_id')}")
            for s in stores_ctx
            if str(s.get("platform") or "") in {"yandex_market", "ozon"}
            and str(s.get("store_id") or "").strip()
        ]
        if not target_uids:
            return JSONResponse({"ok": False, "message": "Нет магазинов для применения"}, status_code=400)
    else:
        target_uids = [f"{platform_id}:{store}"]

    total_upserts = 0
    for uid in target_uids:
        total_upserts += upsert_pricing_logistics_product_settings_bulk(store_uid=uid, rows=rows_to_apply)

    return {
        "ok": True,
        "platform": platform_id,
        "store_id": store,
        "apply_scope": scope,
        "target_stores": len(target_uids),
        "rows_in_file": len(rows_to_apply),
        "upserted": total_upserts,
    }


@router.post("/api/pricing/settings/category-settings/apply-defaults")
async def pricing_settings_apply_defaults(payload: dict):
    platform_id = str(payload.get("platform") or "").strip().lower()
    store_id = str(payload.get("store_id") or "").strip()
    if not platform_id or not store_id:
        return JSONResponse({"ok": False, "message": "platform и store_id обязательны"}, status_code=400)
    store_uid = f"{platform_id}:{store_id}"
    tree = get_pricing_category_tree(store_uid=store_uid)
    dataset_key = str(tree.get("dataset_key") or "").strip()
    if not dataset_key:
        return JSONResponse({"ok": False, "message": "Не найден dataset pricing_categories для выбранного магазина"}, status_code=404)
    try:
        updated = bulk_apply_pricing_defaults(
            dataset_key=dataset_key,
            store_uid=store_uid,
            commission_percent=payload.get("commission_percent"),
            target_margin_percent=payload.get("target_margin_percent"),
            target_margin_rub=payload.get("target_margin_rub"),
            target_profit_rub=payload.get("target_profit_rub"),
            target_profit_percent=payload.get("target_profit_percent"),
            ads_percent=payload.get("ads_percent"),
        )
        _invalidate_pricing_materialized_results_for_store(store_uid=store_uid)
        _invalidate_pricing_read_caches()
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"Не удалось применить значения по умолчанию: {e}"}, status_code=500)
    return {"ok": True, "dataset_key": dataset_key, "updated_rows": updated}


async def _build_pricing_fx_rates_response(*, period: str, date_from: str | None, date_to: str | None) -> dict:
    try:
        today = datetime.date.today()
        p = str(period or "7d").strip().lower()
        if p == "custom":
            if not date_from or not date_to:
                return JSONResponse({"ok": False, "message": "Для custom периода нужны date_from и date_to"}, status_code=400)
            start = _parse_date_ymd(date_from)
            end = _parse_date_ymd(date_to)
        elif p == "14d":
            end = today
            start = today - datetime.timedelta(days=13)
        elif p == "30d":
            end = today
            start = today - datetime.timedelta(days=29)
        else:
            p = "7d"
            end = today
            start = today - datetime.timedelta(days=6)
        if start > end:
            start, end = end, start

        def _fill_forward(rows: list[dict], range_start: datetime.date, range_end: datetime.date) -> list[dict]:
            by_date: dict[str, float] = {}
            for row in rows or []:
                d = str(row.get("date") or "").strip()
                try:
                    r = float(row.get("rate"))
                except Exception:
                    continue
                if d:
                    by_date[d] = r
            out: list[dict] = []
            current = range_start
            last_rate: float | None = None
            while current <= range_end:
                key = current.isoformat()
                if key in by_date:
                    last_rate = by_date[key]
                if last_rate is not None:
                    out.append({"date": key, "rate": round(last_rate, 6)})
                current += datetime.timedelta(days=1)
            # For USD stores the business rule is to show the latest published CBR rate
            # even if the effective date is tomorrow. Keep historical dates intact and
            # only override the current range end when we have exactly "tomorrow" in feed.
            next_day_key = (range_end + datetime.timedelta(days=1)).isoformat()
            if next_day_key in by_date:
                end_key = range_end.isoformat()
                next_rate = round(float(by_date[next_day_key]), 6)
                replaced = False
                for row in out:
                    if row.get("date") == end_key:
                        row["rate"] = next_rate
                        replaced = True
                        break
                if not replaced:
                    out.append({"date": end_key, "rate": next_rate})
            out.sort(key=lambda x: x["date"], reverse=True)
            return out

        cbr_rows = _fill_forward(
            await _fetch_cbr_usd_rates(start, end + datetime.timedelta(days=1)),
            start,
            end,
        )
        # Ozon "для услуг" = курс ЦБ РФ; "для продаж" = курс ЦБ РФ * 1.005
        oz_services_rows = [{"date": row["date"], "rate": row["rate"]} for row in cbr_rows]
        oz_sales_rows = [{"date": row["date"], "rate": round(float(row["rate"]) * 1.005, 6)} for row in cbr_rows]

        response = {
            "ok": True,
            "period": p,
            "date_from": start.isoformat(),
            "date_to": end.isoformat(),
            "tables": {
                "cbr": {"label": "Курс ЦБ РФ", "rows": cbr_rows},
                "ozon_sales": {"label": "Ozon: для продаж", "rows": oz_sales_rows},
                "ozon_services": {"label": "Ozon: для услуг", "rows": oz_services_rows},
            },
        }
        upsert_dashboard_snapshot(
            snapshot_name=_PRICING_FX_RATES_SNAPSHOT_NAME,
            cache_key=_pricing_fx_rates_snapshot_key(period=p, date_from=date_from, date_to=date_to),
            response=response,
        )
        return response
    except Exception as e:
        raise RuntimeError(f"Не удалось загрузить курсы валют: {e}") from e


@router.get("/api/pricing/fx-rates")
async def pricing_fx_rates(period: str = "7d", date_from: str | None = None, date_to: str | None = None):
    normalized_period = str(period or "7d").strip().lower()
    try:
        snapshot = get_dashboard_snapshot(
            snapshot_name=_PRICING_FX_RATES_SNAPSHOT_NAME,
            cache_key=_pricing_fx_rates_snapshot_key(period=normalized_period, date_from=date_from, date_to=date_to),
        )
        if isinstance(snapshot, dict) and isinstance(snapshot.get("response"), dict):
            return snapshot["response"]
        return await _build_pricing_fx_rates_response(period=normalized_period, date_from=date_from, date_to=date_to)
    except RuntimeError as exc:
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=502)
